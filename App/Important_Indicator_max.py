import pandas as pd
import csv
import os
from future.backports.datetime import datetime
from openpyxl import load_workbook
from datetime import datetime
import openpyxl
from copy import copy
import threading
import time
import PySimpleGUI as sg
import os


#园区基站数量
YQNODEB_NUM=6420
#分流比起始时间
ARR_NUM=["17","18","19"]

#sheet2
ywl_lst=[]
ywl_dict=dict()
ywl_sort=dict()
ywl_sort_end=[]
avg_ywl_lst_one=dict()
avg_ywl_lst_two=dict()
avg_ywl_lst_three=dict()
avg_ywl_lst_all=dict()
sheet2_lst=[]
#sheet3 Important_Indicator_max.py
sheet3_data_pp=[]
sheet3_data_qw=[]
sheet3_data=[]
sheet3_sort=dict()
excel_datetime=''
excel_datetime1=''
pp_ll=dict()
pp_sc=dict()
pp_cqsc=dict()
pp_ll_sort=dict()
pp_sc_sort=dict()
pp_cqsc_sort=dict()

qw_ll=dict()
qw_sc=dict()
qw_cqsc=dict()
qw_ll_sort=dict()
qw_sc_sort=dict()
qw_cqsc_sort=dict()

#sheet4
sheet4_data=[]
sheet4_data_pre=[]
sheet4_data_one_4=[]
sheet4_data_one_5=[]
vll_gj_dict_4=dict()
vll_gj_dict_5=dict()
city_lst=['常熟','姑苏','昆山','太仓','吴江','吴中','相城','新区','园区','张家港']
city_ll_4={'常熟':0,'姑苏':0,'昆山':0,'太仓':0,'吴江':0,'吴中':0,'相城':0,'新区':0,'园区':0,'张家港':0}
city_ll_5={'常熟':0,'姑苏':0,'昆山':0,'太仓':0,'吴江':0,'吴中':0,'相城':0,'新区':0,'园区':0,'张家港':0}

#驻留
zl_vll_gj_dict_4=dict()
zl_vll_gj_dict_5=dict()
sheet4_data_two_4=[]
sheet4_data_two_5=[]
sheet4_data_three_4=[]
sheet4_data_three_5=[]
zl_city_lst=['常熟','姑苏','昆山','太仓','吴江','吴中','相城','新区','园区','张家港']
zl_city_ll_4={'常熟':0,'姑苏':0,'昆山':0,'太仓':0,'吴江':0,'吴中':0,'相城':0,'新区':0,'园区':0,'张家港':0}
zl_city_sc_4={'常熟':0,'姑苏':0,'昆山':0,'太仓':0,'吴江':0,'吴中':0,'相城':0,'新区':0,'园区':0,'张家港':0}
zl_city_ll_5={'常熟':0,'姑苏':0,'昆山':0,'太仓':0,'吴江':0,'吴中':0,'相城':0,'新区':0,'园区':0,'张家港':0}
zl_city_sc_5={'常熟':0,'姑苏':0,'昆山':0,'太仓':0,'吴江':0,'吴中':0,'相城':0,'新区':0,'园区':0,'张家港':0}
zl_city_ll_pp_4={'常熟':0,'姑苏':0,'昆山':0,'太仓':0,'吴江':0,'吴中':0,'相城':0,'新区':0,'园区':0,'张家港':0}
zl_city_sc_pp_4={'常熟':0,'姑苏':0,'昆山':0,'太仓':0,'吴江':0,'吴中':0,'相城':0,'新区':0,'园区':0,'张家港':0}
zl_city_ll_pp_5={'常熟':0,'姑苏':0,'昆山':0,'太仓':0,'吴江':0,'吴中':0,'相城':0,'新区':0,'园区':0,'张家港':0}
zl_city_sc_pp_5={'常熟':0,'姑苏':0,'昆山':0,'太仓':0,'吴江':0,'吴中':0,'相城':0,'新区':0,'园区':0,'张家港':0}

#结构质差
jg_city={'常熟':[],'姑苏':[],'昆山':[],'太仓':[],'吴江':[],'吴中':[],'相城':[],'新区':[],'园区':[],'张家港':[]}

#MR覆盖
mr_city={'常熟':[0,0],'姑苏':[0,0],'昆山':[0,0],'太仓':[0,0],'吴江':[0,0],'吴中':[0,0],'相城':[0,0],'新区':[0,0],'园区':[0,0],'张家港':[0,0]}

def read_ywl(file_path):
    df=pd.read_excel(file_path)
    ind=1
    for li in df.values[1:]:
        ywl_dict[li[0]]=[li[1],li[4],li[6]]
        ywl_lst.append([li[0],li[1],li[4],li[6]])
    for li in ywl_lst:
        ywl_sort[li[0]]=li[3]
    ywl_sort_end=sort_dict(ywl_sort)

    for val in ywl_sort_end:
        #缺一个时间
        sheet2_lst.append([val]+ywl_dict[val]+[ind])
        ind+=1

def read_prethreedays_data(file_path):
    df=pd.read_excel(file_path,sheet_name='5G分流比')

    for li in df.values[1:]:
        if str(li[0]) == f'2025-01-{ARR_NUM[0]} 00:00:00':
            avg_ywl_lst_one[li[1]]=[li[1],li[2],li[3]]
        if str(li[0]) == f'2025-01-{ARR_NUM[1]} 00:00:00':
            avg_ywl_lst_two[li[1]]=[li[1],li[2],li[3]]
        if str(li[0]) == f'2025-01-{ARR_NUM[2]} 00:00:00':
            avg_ywl_lst_three[li[1]]=[li[1],li[2],li[3]]

    for li in avg_ywl_lst_one:
        avg_ywl_lst_all[li]=[(avg_ywl_lst_one[li][1]+avg_ywl_lst_two[li][1]+avg_ywl_lst_three[li][1])/3,
                             (avg_ywl_lst_one[li][2]+avg_ywl_lst_two[li][2]+avg_ywl_lst_three[li][2])/3]
    for li in sheet2_lst:
        li.append(avg_ywl_lst_all[li[0]][0])
        li.append(avg_ywl_lst_all[li[0]][1])

def read_sheet3_data(file_path,type):
    global excel_datetime
    global excel_datetime1
    global pp_ll
    global pp_sc
    global pp_cqsc
    global pp_ll_sort
    global pp_sc_sort
    global pp_cqsc_sort

    global qw_ll
    global qw_sc
    global qw_cqsc
    global qw_ll_sort
    global qw_sc_sort
    global qw_cqsc_sort
    try:
        if type=='pp':
            df = pd.read_csv(file_path)
            original_time_str = str(df.values[0][0]).split(' ')[0]
            datetime_obj = datetime.strptime(original_time_str, '%Y-%m-%d')
            year = datetime_obj.year
            month = datetime_obj.month
            day = datetime_obj.day
            excel_datetime = str(year) + '/' + str(month) + '/' + str(day+1)
            excel_datetime1 = str(year) + '/' + str(month) + '/' + str(day)

            for li in df.values:
                sheet3_sort[li[2]] = [li[9], li[10], li[12]]
                pp_ll[li[2]] = li[9]
                pp_sc[li[2]] = li[10]
                pp_cqsc[li[2]] = li[12]
            pp_ll_sort = sort_dict(pp_ll)
            pp_sc_sort = sort_dict(pp_sc)
            pp_cqsc_sort = sort_dict(pp_cqsc)

        elif type=='qw':
            df = pd.read_csv(file_path)
            #excel_datetime = df.values[0][0]
            for li in df.values:
                sheet3_sort[li[2]] = [li[9], li[10], li[12]]
                qw_ll[li[2]] = li[9]
                qw_sc[li[2]] = li[10]
                qw_cqsc[li[2]] = li[12]
            qw_ll_sort = sort_dict(qw_ll)
            qw_sc_sort = sort_dict(qw_sc)
            qw_cqsc_sort = sort_dict(qw_cqsc)

    except FileNotFoundError:
        print(f"文件 {file_path} 未找到，请检查文件路径是否正确。")
    except Exception as e:
        print(f"读取文件 {file_path} 时发生错误: {e}")

def sheet3_data_all():
    for li in pp_ll_sort:
        sheet3_data.append(
            [excel_datetime1, li, pp_ll_sort[li][1], pp_ll_sort[li][0], pp_sc_sort[li][1], pp_sc_sort[li][0],
             pp_cqsc_sort[li][1], pp_cqsc_sort[li][0],
             qw_ll_sort[li][1], qw_ll_sort[li][0], qw_sc_sort[li][1], qw_sc_sort[li][0],
             qw_cqsc_sort[li][1], qw_cqsc_sort[li][0]])
    for li in sheet2_lst:
        li.insert(0, excel_datetime)

def sheet4_data_one(file_path,type):
    with open(file_path, mode='r') as file:
        csv_reader = csv.reader(file)
        if type=='4G':
            for row in csv_reader:
                try:
                    sheet4_data_one_4.append([row[3], row[40], vll_gj_dict_4[row[3]]])
                except:
                    pass
        elif type=='5G':
            for row in csv_reader:
                try:
                    sheet4_data_one_5.append([row[4], row[40], vll_gj_dict_5[row[4]]])
                except:
                    pass

def sheet4_data_one_end_4(type):
    if type=='4G':
        for li in sheet4_data_one_4[1:]:
            city_ll_4[li[2]] += float(li[1])
    elif type=='5G':
        for li in sheet4_data_one_5[1:]:
            city_ll_5[li[2]] += float(li[1])
    #print(city_ll_5)

def sheet4_data_end():
    global sheet4_data
    sum_4=0
    sum_5=0
    avg_rank=0
    avg_sc=0
    avg_ll=0
    avg_sc_pp = 0
    avg_ll_pp = 0
    avg_jg=0
    avg_mr=0
    num=1073741824
    for li in city_lst:
        sheet4_data_pre.append([excel_datetime,li,city_ll_5[li]/num,city_ll_4[li]/num,city_ll_5[li]/(city_ll_4[li]+city_ll_5[li]),
                                zl_city_sc_5[li]/(zl_city_sc_5[li]+zl_city_sc_4[li]),zl_city_ll_5[li]/(zl_city_ll_4[li]+zl_city_ll_5[li]),
                                zl_city_sc_pp_5[li]/(zl_city_sc_pp_5[li]+zl_city_sc_pp_4[li]),zl_city_ll_pp_5[li]/(zl_city_ll_pp_4[li]+zl_city_ll_pp_5[li]),' ',' ',
                                jg_city[li][1]/jg_city[li][0],' ',mr_city[li][0]/mr_city[li][1],' '])
    sheet4_data+=sheet4_data_pre
    for li in sheet4_data_pre:
        sum_5+=li[2]
        sum_4+=li[3]
        avg_rank+=li[4]
        avg_sc+=li[5]
        avg_ll+=li[6]
        avg_sc_pp += li[7]
        avg_ll_pp += li[8]
        avg_jg += li[11]
        avg_mr += li[13]
    sheet4_data.append([excel_datetime,'总计',sum_5,sum_4,avg_rank/10,avg_sc/10,avg_ll/10,avg_sc_pp/10,avg_ll_pp/10,' ',' ',avg_jg,' ',avg_mr/10,' '])

def read_fl_gj(file_path):
    pf=pd.read_excel(file_path,sheet_name='4G')
    pf2=pd.read_excel(file_path,sheet_name='5G')
    for li in pf.values:
        vll_gj_dict_4[li[0]]=li[1]
    for li in pf2.values:
        vll_gj_dict_5[li[0]]=li[1]

def sheet4_data_two(file_path,type,type2):
    pf=pd.read_csv(file_path)
    if type == '4G':
        for row in pf.values:
            try:
                if str(row[9])!='nan' and str(row[10])!='nan' and type2=='qw':
                   sheet4_data_two_4.append([row[2], row[9], row[10], zl_vll_gj_dict_4[row[2]]])
                elif str(row[9])!='nan' and str(row[10])!='nan' and type2=='pp':
                    sheet4_data_three_4.append([row[2], row[9], row[10], zl_vll_gj_dict_4[row[2]]])
            except:
                pass
    elif type == '5G':
        for row in pf.values:
            try:
                if str(row[9])!='nan' and str(row[10])!='nan' and type2=='qw':
                    sheet4_data_two_5.append([row[2], row[9], row[10], zl_vll_gj_dict_5[row[2]]])
                elif str(row[9])!='nan' and str(row[10])!='nan' and type2=='pp':
                    sheet4_data_three_5.append([row[2], row[9], row[10], zl_vll_gj_dict_5[row[2]]])
            except:
                pass

def sheet4_data_two_end(type):
    global sheet4_data
    if type=='qw':
        for li in sheet4_data_two_4:
            zl_city_ll_4[li[3]]+=li[1]
            zl_city_sc_4[li[3]]+=li[2]
        for li in sheet4_data_two_5:
            zl_city_ll_5[li[3]]+=li[1]
            zl_city_sc_5[li[3]]+=li[2]
    elif type=='pp':
        for li in sheet4_data_three_4:
            zl_city_ll_pp_4[li[3]]+=li[1]
            zl_city_sc_pp_4[li[3]]+=li[2]
        for li in sheet4_data_three_5:
            zl_city_ll_pp_5[li[3]]+=li[1]
            zl_city_sc_pp_5[li[3]]+=li[2]

def read_zl_gj(file_path):
    pf = pd.read_excel(file_path, sheet_name='4G')
    pf2 = pd.read_excel(file_path, sheet_name='5G')
    for li in pf.values:
        zl_vll_gj_dict_4[li[0]] = li[2]
    for li in pf2.values:
        zl_vll_gj_dict_5[li[0]] = li[2]

def read_jg(file_path):
    try:
        with open(file_path, mode='r') as file:
            csv_reader = csv.reader(file)
            for row in list(csv_reader)[1:]:
                city=row[5][:2]
                if city=='吴中':
                    jg_city[city] += [int(row[6])-YQNODEB_NUM, 0]
                    jg_city['园区'] += [YQNODEB_NUM, int(row[7])]
                    continue
                if city=='张家':
                    city='张家港'
                if city=='虎丘':
                    city='新区'
                jg_city[city]+=[int(row[6]),int(row[7])]
            #print(jg_city)

    except FileNotFoundError:
        print(f"文件 {file_path} 未找到，请检查文件路径是否正确。")
    except Exception as e:
        print(f"读取文件时发生错误: {e}")

def read_mr(file_path):
    try:
        with open(file_path, mode='r') as file:
            csv_reader = csv.reader(file)
            for row in list(csv_reader)[1:]:
                #print(row[5][:2])
                city=row[5][:2]
                if city=='虎丘':
                    city='新区'
                if city=='张家':
                    city='张家港'
                if city=='吴中' and '园区' in row[8]:
                    city='园区'
                #print(city)
                mr_city[city][0]+=float(row[14])*float(row[19])/100
                mr_city[city][1]+=float(row[14])

            #print(mr_city)


    except FileNotFoundError:
        print(f"文件 {file_path} 未找到，请检查文件路径是否正确。")
    except Exception as e:
        print(f"读取文件时发生错误: {e}")

def sort_dict(input_dict):
    # 将字典的键值对转换为元组列表，元组包含 (键, 值)
    items_list = [(key, value) for key, value in input_dict.items()]
    # 对元组列表进行降序排序，根据值排序
    sorted_items = sorted(items_list, key=lambda x: x[1], reverse=True)
    result_dict = {}
    for index, (key, value) in enumerate(sorted_items):
        result_dict[key] = [index+1, value]
    return result_dict

#添加数据到excel
def append_data_to_excel(file_path, sheet_name,sheet_name2,sheet_name3, data,data2,data3):
    # 加载现有的 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表
    sheet = workbook[sheet_name]
    sheet2 = workbook[sheet_name2]
    sheet3 = workbook[sheet_name3]
    # 遍历要追加的数据列表
    for row in data:
        # 将数据逐行添加到工作表的下一行
        sheet.append(row)
    for row in data2:
        # 将数据逐行添加到工作表的下一行
        sheet2.append(row)
    for row in data3:
        # 将数据逐行添加到工作表的下一行
        sheet3.append(row)
    # 保存更新后的 Excel 文件
    workbook.save(file_path)

#添加样式
def get_second_row_styles(excel_file,num, sheet_name):
    # 加载 Excel 文件
    workbook = load_workbook(excel_file)
    # 选择指定的工作表
    sheet = workbook[sheet_name]
    second_row_styles = []
    # 遍历第二行的每个单元格
    for cell in sheet[num]:
        # 复制 StyleProxy 对象
        font = copy(cell.font)
        border = copy(cell.border)
        fill = copy(cell.fill)
        number_format = copy(cell.number_format)
        alignment = copy(cell.alignment)
        # 存储复制后的 StyleProxy 对象的元组
        style = (font, border, fill, number_format, alignment)
        second_row_styles.append(style)
    return second_row_styles


def apply_styles_to_last_two_rows(excel_file,num, sheet_name):
    # 获取第二行的样式
    second_row_styles = get_second_row_styles(excel_file,num, sheet_name)
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]
    # 获取最后两行的行号
    max_row = sheet.max_row
    last_two_rows = [max_row-12,max_row-11,max_row-10,max_row-9,max_row-8,max_row-7,max_row-6,max_row-5,max_row-4,max_row-3,max_row-2,max_row-1, max_row]
    # 应用样式到最后几行
    # 应用样式到最后两行
    for row_num in last_two_rows:
        for i, style in enumerate(second_row_styles):
            new_cell = sheet.cell(row=row_num, column=i + 1)
            new_cell.font, new_cell.border, new_cell.fill, new_cell.number_format, new_cell.alignment = style
    # 保存修改后的 Excel 文件
    workbook.save(excel_file)



if __name__ == '__main__':

    # path_str1 = 'C:/Users/24253/Desktop/工作内容/25年每月例行工作/1月工作内容/指标提取/'
    # path_str2 = 'C:/Users/24253/Desktop/工作内容/25年每月例行工作/1月工作内容/'
    # path_str3 = 'C:/Users/24253/Desktop/工作内容/学习资料/江苏格安/工具/工具/'

    #max_excel_path = f'{path_str2}5G重点指标通报-0109.xlsx'
    #update_excel_path = f'{path_str2}5G重点指标通报-0108.xlsx'
    # fl_gj_path = f'{path_str3}分区县V流量.xlsx'
    # zl_gj_path = f'{path_str3}驻留小区区县V驻留最新.xlsx'
    #fl_gj_path = '分区县V流量.xlsx'
    #zl_gj_path = '驻留小区区县V驻留最新.xlsx'

    layout = [[sg.Text("选择文件 5G服务小区覆盖分析--小区详情数据:")],
              [sg.InputText(key='-FILE_A-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 5G终端4G小区流量分析-全网:")],
              [sg.InputText(key='-FILE_B-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 5G终端5G小区流量分析-全网:")],
              [sg.InputText(key='-FILE_C-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 5G终端4G小区流量分析-品牌:")],
              [sg.InputText(key='-FILE_D-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 5G终端5G小区流量分析-品牌:")],
              [sg.InputText(key='-FILE_E-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 5G终端地市分析列表-全网:")],
              [sg.InputText(key='-FILE_F-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 5G终端地市分析列表-品牌:")],
              [sg.InputText(key='-FILE_G-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 分流比4G:")],
              [sg.InputText(key='-FILE_H-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 分流比5G:")],
              [sg.InputText(key='-FILE_I-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 小区网络结构分析--地市全区数据:")],
              [sg.InputText(key='-FILE_J-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 业务类:")],
              [sg.InputText(key='-FILE_K-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 分区县V流量:")],
              [sg.InputText(key='-FILE_M-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 驻留小区区县V驻留:")],
              [sg.InputText(key='-FILE_N-', enable_events=True), sg.FilesBrowse()],
              [sg.Text("选择文件 重点指标文件:")],
              [sg.InputText(key='-FILE_L-', enable_events=True), sg.FilesBrowse()],
              [sg.Button('提交')]]

    window = sg.Window('选择所需文件', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSE_ATTEMPTED_EVENT or event == 'Cancel':
            break
        if event == '提交':
            sheet4_path_mr = values['-FILE_A-']
            sheet4_path_zl_4 = values['-FILE_B-']
            sheet4_path_zl_5 = values['-FILE_C-']
            sheet4_path_zl_pp_4 = values['-FILE_D-']
            sheet4_path_zl_pp_5 = values['-FILE_E-']
            sheet3_path_qw = values['-FILE_F-']
            sheet3_path_pp = values['-FILE_G-']
            sheet4_path_fl_4 = values['-FILE_H-']
            sheet4_path_fl_5 = values['-FILE_I-']
            sheet4_path_jg = values['-FILE_J-']
            ywl_path = values['-FILE_K-']
            update_excel_path = values['-FILE_L-']
            fl_gj_path = values['-FILE_M-']
            zl_gj_path = values['-FILE_N-']

            #print(path_str1)

            window.close()
        if event == sg.WIN_CLOSED:
            break

    read_ywl(ywl_path)
    read_prethreedays_data(update_excel_path)



    read_sheet3_data(sheet3_path_pp,type='pp')
    read_sheet3_data(sheet3_path_qw,type='qw')
    sheet3_data_all()
    read_fl_gj(fl_gj_path)

    sheet4_data_one(sheet4_path_fl_4,'4G')
    sheet4_data_one(sheet4_path_fl_5,'5G')
    sheet4_data_one_end_4('4G')
    sheet4_data_one_end_4('5G')
    read_zl_gj(zl_gj_path)
    sheet4_data_two(sheet4_path_zl_4,'4G','qw')
    sheet4_data_two(sheet4_path_zl_5,'5G','qw')
    sheet4_data_two_end('qw')
    sheet4_data_two(sheet4_path_zl_pp_4, '4G', 'pp')
    sheet4_data_two(sheet4_path_zl_pp_5, '5G', 'pp')
    sheet4_data_two_end('pp')
    read_jg(sheet4_path_jg)
    read_mr(sheet4_path_mr)

    #
    sheet4_data_end()


    append_data_to_excel(update_excel_path,'5G分流比','5G驻留比','分区县',sheet2_lst,sheet3_data,sheet4_data)
    # 应用样式到最后两行
    apply_styles_to_last_two_rows(update_excel_path, 16240, '5G分流比')
    apply_styles_to_last_two_rows(update_excel_path, 15588, '5G驻留比')
    apply_styles_to_last_two_rows(update_excel_path, 11880, '分区县')
