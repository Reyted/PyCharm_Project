import pandas as pd
from openpyxl import Workbook, load_workbook
import time

def read_csv_with_encoding(file_path, encodings=['utf-8', 'gb2312', 'gbk', 'gb18030']):
    for encoding in encodings:
        try:
            return pd.read_csv(file_path, encoding=encoding, dtype=str)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Unable to read the file {file_path} with any of the specified encodings.")

def write_excel_file(data, filename):
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = '基带板-全网'
    sheet2 = wb.create_sheet("RRU-全网")

    sheet2.append(['网元名称','单板名称','单板类型','生产日期','机柜号','机框号','槽位号','特殊信息','资产序列号','网元型号','小区名称'])
    sheet1.append(['网元名称','单板名称','单板类型','生产日期','机柜号','机框号','槽位号','特殊信息','资产序列号'])

    for row in data:
        row = [str(item) for item in row]
        if row[1] in 'UBBPLBBP':
            sheet1.append([row[0],row[1],row[2],row[3],row[6],row[4],row[8],row[5],row[7]])
        elif row[1] in 'MRRUAIRUMRFUMPMURRN':
            bol = any(x in row[5] for x in ['RRU5', 'AAU5', 'RRU3', 'RRU7', 'AAU3', 'RRN3'])
            if not bol:
                sheet2.append([row[0], row[1], row[2], row[3], row[6], row[4], row[8], row[5], row[7], ' '])
            else:
                li = row[5].split(',')
                if len(li) == 1 or not bol:
                    sheet2.append([row[0],row[1],row[2],row[3],row[6],row[4],row[8],row[5],row[7],' '])
                elif len(li) > 1 and bol:
                    for i in li:
                        li2 = i.split(' ')
                        for j in li2:
                            if any(x in j for x in ['RRU5', 'AAU5', 'RRU3', 'RRU7', 'AAU3', 'RRN3']):
                                sheet2.append([row[0],row[1],row[2],row[3],row[6],row[4],row[8],row[5],row[7],j.split('(')[0]])
                                break
                        else:
                            continue
                        break

    today = time.strftime("%Y-%m-%d", time.localtime())
    wb.save(filename + f'硬件信息{today}.xlsx')

def lst_cell(filepath1, filepath2, url):
    df1 = read_csv_with_encoding(filepath1)
    df2 = read_csv_with_encoding(filepath2)

    df = pd.merge(df1, df2, left_on=[df1.columns[1], df1.columns[5]], right_on=[df2.columns[1], df2.columns[4]])
    df['小区名称'] = df[df2.columns[5]]
    df = df.drop(columns=[df2.columns[5]])

    df.to_csv(f'{url}test.csv', index=False, encoding='utf-8-sig')

def safe_split(s, sep, index, default=''):
    parts = s.split(sep)
    return parts[index] if index < len(parts) else default

def merge_cell(filepath1, filepath2, name):
    df1 = read_csv_with_encoding(filepath1)
    df2 = pd.read_excel(filepath2, sheet_name=name)

    if name == 'RRU-全网':
        df1['indefy'] = df1.apply(lambda row: row[1] + safe_split(str(row[6]), '-', 0) + '-' +
                                               safe_split(str(row[6]), '-', 1) + '-' +
                                               safe_split(str(row[6]), '-', 2), axis=1)
        df2['indefy'] = df2.apply(lambda row: str(row[0]) + str(row[4]) + '-' + str(row[5]) + '-' + str(row[6]), axis=1)
    else:
        df1['indefy'] = df1.apply(lambda row: row[1] + safe_split(str(row[8]), '-', 0) + '-' +
                                               safe_split(str(row[8]), '-', 1) + '-' +
                                               safe_split(str(row[8]), '-', 2), axis=1)
        df2['indefy'] = df2.apply(lambda row: str(row[0]) + str(row[4]) + '-' + str(row[5]) + '-' + str(row[6]), axis=1)

    df_merged = df2[~df2['indefy'].isin(df1['indefy'])]

    with pd.ExcelWriter(filepath2, engine='openpyxl', mode='a') as writer:
        df_merged.to_excel(writer, sheet_name=f'{name.split("-")[0]}-冗余', index=False)

def second_filter(filepath1, filepath2, name):
    df1 = pd.read_excel(filepath1, sheet_name=name)
    df2 = read_csv_with_encoding(filepath2)

    df_filtered = df1[df1['网元名称'].isin(df2['网元名称'])]

    with pd.ExcelWriter(filepath1, engine='openpyxl', mode='a') as writer:
        df_filtered.to_excel(writer, sheet_name=name, index=False)

def merge_csv(file1, file2, output_file):
    df1 = read_csv_with_encoding(file1)
    df2 = read_csv_with_encoding(file2)

    columns = ['网元名称', '单板名称', '单板类型', '生产日期', '机柜号', '机框号', '槽位号', '特殊信息', '资产序列号']
    df_merged = pd.concat([df1[columns], df2[columns]], ignore_index=True)

    df_merged.to_csv(output_file, index=False, encoding='utf-8-sig')

if __name__ == '__main__':
    url = 'C:/Users/24253/Desktop/Python/PLAY/'
    today = time.strftime("%Y-%m-%d", time.localtime())

    merge_csv(f'{url}存量_板_20241017_114851.csv', f'{url}存量_板_20241017_115016.csv', f'{url}merged_file.csv')

    df = read_csv_with_encoding(f'{url}merged_file.csv')
    write_excel_file(df.values, url)

    lst_cell(f'{url}查询小区物理单板拓扑关系.csv', f'{url}查询小区静态参数.csv', url)

    merge_cell(f'{url}查询小区物理单板拓扑关系.csv', f'{url}硬件信息{today}.xlsx', '基带板-全网')
    second_filter(f'{url}硬件信息{today}.xlsx', f'{url}查询小区物理单板拓扑关系.csv', '基带板-冗余')

    merge_cell(f'{url}查询小区物理单板拓扑关系.csv', f'{url}硬件信息{today}.xlsx', 'RRU-全网')
    second_filter(f'{url}硬件信息{today}.xlsx', f'{url}查询小区物理单板拓扑关系.csv', 'RRU-冗余')