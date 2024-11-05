import csv
from multiprocessing.pool import ThreadPool

import openpyxl
import time
import pandas as pd
from openpyxl import load_workbook
import threading


#写数据
def write_csv_file(data, filename):
    # 创建一个新的Excel工作簿
    workbook = openpyxl.Workbook()

    #当前活动Sheet
    sheet1 = workbook.active
    sheet1.title = '基带板-全网'
    # 创建一个新的表单
    sheet2 = workbook.create_sheet("RRU-全网")

    #添加一行数据
    sheet2.append(['网元名称','单板名称','单板类型','生产日期','机柜号','机框号','槽位号','特殊信息','资产序列号','网元型号','小区名称'])
    sheet1.append(['网元名称','单板名称','单板类型','生产日期','机柜号','机框号','槽位号','特殊信息','资产序列号'])


    for row in data:
        bol=('RRU5' in row[5] or 'AAU5' in row[5] or 'RRU3' in row[5] or 'RRU7' in row[5] or 'AAU3' in row[5] or 'RRN3' in row[5])
        li=[row[0],row[1],row[2],row[3],row[6],row[4],row[8],row[5],row[7]]
        if row[1] in 'UBBPLBBP':
            sheet1.append([row[0],row[1],row[2],row[3],row[6],row[4],row[8],row[5],row[7]])
        if row[1] in 'MRRUAIRUMRFUMPMURRN':

            if not bol:
                sheet2.append([row[0], row[1], row[2], row[3], row[6], row[4], row[8], row[5], row[7], ' '])
            else:
                li = row[5].split(',')
                if  len(li)==1 or not bol:
                    sheet2.append([row[0],row[1],row[2],row[3],row[6],row[4],row[8],row[5],row[7],' '])
                elif len(li)>1 and bol:
                    for i in li:
                        li2=i.split(' ')
                        if len(li2)==1:
                            if 'RRU5' in i or 'AAU5' in i or 'RRU3' in i or 'RRU7' in i or 'AAU3' in i or 'RRN3' in i:
                                sheet2.append([row[0],row[1],row[2],row[3],row[6],row[4],row[8],row[5],row[7],i.split('(')[0]])
                                break
                        else:
                            for j in li2:
                                if 'RRU5' in j or 'AAU5' in j or 'RRU3' in j or 'RRU7' in j or 'AAU3' in j or 'RRN3' in j:
                                    sheet2.append([row[0],row[1],row[2],row[3],row[6],row[4],row[8],row[5],row[7],j.split('(')[0]])
                                    break

    # 保存工作簿
    #获取当前时间
    today = time.strftime("%Y-%m-%d", time.localtime())
    workbook.save(filename+f'硬件信息{today}.xlsx')


#小区查询
def lst_cell(filepath1:str,filepath2:str,url:str):
    lst = []
    f1lst=[]
    f2lst=[]
    celllst=[]

    def open_filepath1(filepath:str):
        with open(filepath, 'r') as f:
            reader = csv.reader(f)
            for row in reader:
                f1lst.append(row)

    def open_filepath2(filepath:str):
        with open(filepath, 'r') as f:
            reader = csv.reader(f)
            for row in reader:
                f2lst.append(row)

    #process1=mp.Process(target=open_filepath1, args=(filepath1,)).start()
    #threading.Thread(target=open_filepath1, args=(filepath1,)).start()
    #threading.Thread(target=open_filepath2, args=(filepath2,)).start()

    #process2=mp.Process(target=open_filepath2, args=open_filepath2).start()

    with open(filepath1, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            f1lst.append(row)

    with open(filepath2, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            f2lst.append(row)

    for row in f1lst:
        indentify = row[1] + row[5]
        for row1 in f2lst:
            indentify1 = row1[1] + row1[4]
            if indentify == indentify1:
                celllst.append(row1[5])

    data = []
    ind = 0

    with open(filepath1, 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            data.append(row)

    for cell in celllst:
        data[ind].append(cell)
        ind += 1

    with open(f'{url}test.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(data)

#小区合并 冗余
def merge_cell(filepath1:str,filepath2:str,name:str):
    lstt = []
    lst1 = []
    lst2 = []
    lst3 = ''
    teststr = ''

    with open(filepath1, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            lst1.append(row)
    if name=='RRU-全网':
        for li in lst1:
            lst = li[6].split('-')
            if len(lst) > 1:
                teststr += li[1] + str(lst[0]) + '-' + str(lst[1]) + '-' + str(lst[2])
            else:
                lst1=li[6].split('/')
                if len(lst1)>1:
                    teststr += li[1] + str(lst1[0][-1]) + '-' + str(lst1[1]) + '-' + str(lst1[2])
    else:
        for li in lst1:
            lst = li[8].split('-')
            if len(lst) > 2:
                teststr += li[1] + str(lst[0]) + '-' + str(lst[1]) + '-' + str(lst[2])

    if name=='RRU-全网':
        df = pd.read_excel(filepath2, sheet_name='RRU-全网')
    else:
        df = pd.read_excel(filepath2, sheet_name='基带板-全网')
    workbook = load_workbook(filepath2)

    for row in df.values:
        lst2.append(list(row))
        indefy2 = row[0] + str(row[4]) + '-' + str(row[5]) + '-' + str(row[6])
        if indefy2 in teststr:
            lstt.append('1')
        else:
            lstt.append('2')

    if name=='RRU-全网':
        sheet3 = workbook.create_sheet('RRU-冗余')
        ind1 = 0
        ind2 = 0
        sheet3.append(['网元名称', '单板名称', '单板类型', '生产日期', '机柜号', '机框号', '槽位号', '特殊信息', '资产序列号','网元型号'])
        for li in lstt:
            if li == '2': #and not (lst2[ind1][0] in lst3)
                sheet3.append(lst2[ind1])
            else:
                ind2 += 1
            ind1 += 1
    else:
        sheet4 = workbook.create_sheet('基带板-冗余')
        sheet4.append(
            ['网元名称', '单板名称', '单板类型', '生产日期', '机柜号', '机框号', '槽位号', '特殊信息', '资产序列号'])
        ind1 = 0
        ind2 = 0
        for li in lstt:
            if li == '2':
                sheet4.append(lst2[ind1])
            else:
                ind2 += 1
            ind1 += 1

    # 保存修改后的Excel文件
    workbook.save(filepath2)

def second_filter(filepath1:str,filepath2:str,name:str):
    brd_list=[]
    rru_list=[]
    topo_str=''
    if name=='基带板-冗余':

        with open(filepath2, 'r', newline='') as f:
            reader = csv.reader(f)
            for row in reader:
                topo_str+=row[1]

        df = pd.read_excel(filepath1, sheet_name=name)
        for row in df.values:
            if row[0] in topo_str:
                brd_list.append(list(row))

        wb=load_workbook(filepath1)
        sheet=wb['基带板-冗余']
        wb.remove(sheet)
        sheet2=wb.create_sheet('基带板-冗余')
        sheet2.append(
            ['网元名称', '单板名称', '单板类型', '生产日期', '机柜号', '机框号', '槽位号', '特殊信息', '资产序列号'])
        for li in brd_list:
            sheet2.append(li)
        wb.save(filepath1)
    else:
        with open(filepath2, 'r', newline='') as f:
            reader = csv.reader(f)
            for row in reader:
                topo_str+=row[1]

        df = pd.read_excel(filepath1, sheet_name=name)
        for row in df.values:
            if row[0] in topo_str:
                rru_list.append(list(row))

        wb=load_workbook(filepath1)
        sheet=wb['RRU-冗余']
        wb.remove(sheet)
        sheet2=wb.create_sheet('RRU-冗余')
        sheet2.append(
            ['网元名称', '单板名称', '单板类型', '生产日期', '机柜号', '机框号', '槽位号', '特殊信息', '资产序列号','网元型号'])
        for li in rru_list:
            sheet2.append(li)
        wb.save(filepath1)


#合并函数
def merge_csv(file1, file2, output_file):
    li1 = []
    li2 = []
    #打开第一个CSV文件并读取内容
    with open(file1, 'r') as f1:
        reader1 = csv.reader(f1)
        index=[]
        ind=0
        data1 = list(reader1)
        for i in data1[0]:
            if i in '网元名称单板名称单板类型生产日期机柜号机框号槽位号特殊信息资产序列号':
                index.append(ind)
            ind+=1
        for row in data1:
            li=[]
            for i in index:
                li.append(row[i])
            li1.append(li)

    #打开第二个CSV文件并读取内容
    with open(file2, 'r') as f2:
        reader2 = csv.reader(f2)
        index = []
        ind = 0
        data2 = list(reader2)
        for i in data2[0]:
            if i in '网元名称单板名称单板类型生产日期机柜号机框号槽位号特殊信息资产序列号':
                index.append(ind)
            ind+=1
        for row in data2:
            li = []
            for i in index:
                li.append(row[i])
            li2.append(li)

    # 合并两个CSV文件的内容
    merged_data = li1+li2

    #写入新的CSV文件
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(merged_data)

#筛选函数
def read_csv_file(filepath):
    with open(filepath, 'r') as f:
        reader = csv.reader(f)
        data = []
        for row in reader:
            data.append(row)
        write_csv_file(data, 'C:/Users/24253/Desktop/PLAY/')


if __name__ == '__main__':

    #文件存储路径
    url='C:/Users/24253/Desktop/PLAY/'


    today = time.strftime("%Y-%m-%d", time.localtime())
    # 调用函数进行合并
    merge_csv(f'{url}存量_板_20241017_114851.csv', f'{url}存量_板_20241017_115016.csv', f'{url}merged_file.csv')

    read_csv_file(f'{url}merged_file.csv')

    lst_cell(f'{url}查询小区物理单板拓扑关系.csv',f'{url}查询小区静态参数.csv',url)

    #threading.Thread(target=merge_cell,args=(f'{url}查询小区物理单板拓扑关系.csv',f'{url}硬件信息{today}.xlsx', '基带板-全网')).start()
    #threading.Thread(target=merge_cell,args=(f'{url}查询小区物理单板拓扑关系.csv',f'{url}硬件信息{today}.xlsx', 'RRU-全网')).start()


    # 基带板冗余
    merge_cell(f'{url}查询小区物理单板拓扑关系.csv',f'{url}硬件信息{today}.xlsx', '基带板-全网')
    second_filter(f'{url}硬件信息{today}.xlsx',f'{url}查询小区物理单板拓扑关系.csv','基带板-冗余')


    # RRU冗余
    merge_cell(f'{url}查询小区物理单板拓扑关系.csv', f'{url}硬件信息{today}.xlsx','RRU-全网')
    second_filter(f'{url}硬件信息{today}.xlsx',f'{url}查询小区物理单板拓扑关系.csv', 'RRU-冗余')