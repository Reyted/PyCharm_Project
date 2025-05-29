import pandas as pd
from numpy.ma.core import append
import openpyxl
from openpyxl import load_workbook
import csv
import os
from future.backports.datetime import datetime
from datetime import datetime
from copy import copy
import time
import os

FILE_TIME='0106-0112'
TXS={'0':[],'1':[],'2':[],'3':[],'4':[],'5':[],'6':[],'7':[]}
ECGI={}
MML=[]
# mml站名-扇区设备编号-端口号
MML_2={}
# mml站名-柜号-框号-槽号
MML_3={}
# mml站名-扇区编号
MML_4={}
# mml站名-扇区编号/RRU数目
MML_5={}
# 站型&区县
NODE_INFO={}
# RRU型号及串号
RRU_ID={}

def read_rru_id(filepath):
    # 使用 openpyxl 直接加载工作簿
    wb = load_workbook(filename=filepath, read_only=True)
    sheet = wb.active
    # 使用生成器按行读取数据，避免一次性加载所有数据
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    # 转换为 pandas DataFrame
    df = pd.DataFrame(data[1:], columns=data[0])
    # 打印读取的数据
    for li in df.values.tolist():
        # 站名-柜框槽
        str1=li[0]+'-'+li[3]
        RRU_ID[str1]=[li[7],li[8]]

def read_gc(filepath):
    with open(filepath, 'r',encoding='utf-8') as f:
        csv_reader = csv.reader(f, delimiter=',')
        for row in csv_reader:
            ECGI[row[1]+'-'+row[17]]=row[35]
            try:
                try:
                    if NODE_INFO[row[1]] == []:
                        NODE_INFO[row[1]] = [row[4], row[7]]
                except:
                    NODE_INFO[row[1]] = []
                    NODE_INFO[row[1]] = [row[4], row[7]]
            except:
                pass

def read_mml(filepath):
    pf=pd.read_excel(filepath,sheet_name='查询扇区设备天线配置信息')
    for li in pf.values:
        str5=li[1]+'-'+str(li[5])+'-'+str(li[6])+'-'+str(li[7])
        str6=''
        try:
            str6=RRU_ID[str5][1]
        except:
            str6='-'
        MML.append([li[1],li[4],li[5],li[6],li[7],li[8],str6])
    old_id=''
    old_val=0
    MML[0].append(0)
    old_val=MML[0][6]
    for li in MML:
        # 站名-扇区设备编号
        str1=li[0]+'-'+str(li[1])
        try:
            try:
                if li[6]=='-':
                    MML_5[str1]=[]
                elif li[6] not in MML_5[str1]:
                    MML_5[str1].append(li[6])
            except:
                MML_5[str1]=[]
                MML_5[str1].append(li[6])
        except:
            MML_5[str1]=[]
        if str1==old_id:
            li.append(old_val+1)
            try:
                MML_4[str1].append(old_val + 1)
            except:
                MML_4[str1]=[]
        else:
            li.append(0)
            try:
                MML_4[str1]=[]
                MML_4[str1].append(0)
            except:
                MML_4[str1]=[]
            #print()
        old_id = str1
        old_val = li[7]
        # 站名-扇区设备编号-端口号
        str2=str1+'-'+str(li[7])
        # 站名-柜号-框号-槽号
        str3=li[0]+'-'+str(li[2])+'-'+str(li[3])+'-'+str(li[4])
        MML_2[str2]=[li[2],li[3],li[4],li[5]]
        try:
            MML_3[str3] = RRU_ID[str3]
        except:
            pass
    #print(MML_2['LE5E9A1-3-0'])
    #print(MML_4)
    #print(MML_5)

def read_temp_selectfile(filepath):
    pf=pd.read_excel(filepath)
    lst=list(pf.values)
    for li in lst:
        little_li(0,li)
        little_li(2,li)
        little_li(4,li)
        little_li(6,li)
        little_li(8,li)
        little_li(10,li)
        little_li(12,li)
        little_li(14,li)
    print(len(TXS['0']))
    print(len(TXS['1']))
    print(len(TXS['2']))
    print(len(TXS['3']))
    print(len(TXS['4']))
    print(len(TXS['5']))
    print(len(TXS['6']))
    print(len(TXS['7']))

def little_li(num,li):
    # 天线通道号
    channel_id=''
    try:
        cz = int(li[num+9]) - int(li[num+10])
        if li[8] > 0.01 and cz < 5:
            # 基站-本地小区标识
            str1 = li[1] + '-' + str(li[4])
            # 基站-扇区设备编号
            str2 = li[1] + '-' + str(li[3])
            # 基站-扇区设备编号-端口号
            str3=str2 + '-' + f'{int(num/2)}'
            # 基站-柜框槽
            try:
                str4 = li[1] + '-' + str(MML_2[str3][0]) + '-' + str(MML_2[str3][1]) + '-' + str(MML_2[str3][2])
            except:
                str4 = li[1] + '---'
            #str4=li[1]+'-'+str(MML_2[str3][0])+'-'+str(MML_2[str3][1])+'-'+str(MML_2[str3][2])
            try:
                lj1=NODE_INFO[li[1]][0]
            except:
                lj1='未查询到'
            try:
                lj2=NODE_INFO[li[1]][1]
            except:
                lj2='未查询到'
            try:
                lj3=len(MML_5[str2])
            except:
                lj3='未查询到'
            try:
                lj4=len(MML_4[str2])
            except:
                lj4='未查询到'
            try:
                lj5=ECGI[str1]
            except:
                lj5='未查询到'
            try:
                lj6=MML_2[str3][1]
            except:
                lj6='未查询到'
            try:
                lj7 = RRU_ID[str4][0]
            except:
                lj7 = '未查询到'
            try:
                lj8 = MML_2[str3][3]
            except:
                lj8 = '未查询到'
            try:
                lj9 = RRU_ID[str4][1]
            except:
                lj9 = '未查询到'
            lj=[FILE_TIME,lj1,lj2,li[1], li[3], li[4],lj3,lj4 ,str1,lj5,lj6,lj7,'-','-' ,lj8,str3, int(num/2), lj9,str2, li[num+9], li[num+10], cz, li[8]]
            # lj=[FILE_TIME,li[1], li[3], li[4],str1,'-','-' ,str2, li[9], li[10], cz, li[8]]
            TXS[f'{int(num/2)}'].append(lj)
    except:
        pass

#添加数据到excel
def append_data_to_excel(file_path, sheet_name, data):
    # 加载现有的 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表
    sheet = workbook[sheet_name]

    # sheet=workbook.create_sheet(sheet_name)
    # 遍历要追加的数据列表
    for row in data:
        # 将数据逐行添加到工作表的下一行
        sheet.append(row)
    # 保存更新后的 Excel 文件
    workbook.save(file_path)

#添加样式
def get_second_row_styles(excel_file,num, sheet_name):
    # 加载 Excel 文件
    workbook = load_workbook(excel_file)
    # 选择指定的工作表
    sheet = workbook[sheet_name]
    second_row_styles = []
    # 遍历第二行的每个单元格
    for cell in sheet[num]:
        # 复制 StyleProxy 对象
        font = copy(cell.font)
        border = copy(cell.border)
        fill = copy(cell.fill)
        number_format = copy(cell.number_format)
        alignment = copy(cell.alignment)
        # 存储复制后的 StyleProxy 对象的元组
        style = (font, border, fill, number_format, alignment)
        second_row_styles.append(style)
    return second_row_styles


def apply_styles_to_last_two_rows(excel_file,num, sheet_name):
    # 获取第二行的样式
    second_row_styles = get_second_row_styles(excel_file,num, sheet_name)
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]
    # 获取最后两行的行号
    max_row = sheet.max_row
    last_two_rows=[]
    for i in range(max_row-2):
        last_two_rows.append(i+3)
    #last_two_rows = [max_row-12,max_row-11,max_row-10,max_row-9,max_row-8,max_row-7,max_row-6,max_row-5,max_row-4,max_row-3,max_row-2,max_row-1, max_row]
    # 应用样式到最后几行
    # 应用样式到最后两行
    for row_num in last_two_rows:
        for i, style in enumerate(second_row_styles):
            new_cell = sheet.cell(row=row_num, column=i + 1)
            new_cell.font, new_cell.border, new_cell.fill, new_cell.number_format, new_cell.alignment = style
    # 保存修改后的 Excel 文件
    workbook.save(excel_file)

if __name__ == '__main__':
    temp_select_filepath='G:/工作内容/PRS数据/SECTOREQM/临时查询_查询结果_20250117-泰山.xlsx'
    temp_select_filepath2='G:/工作内容/PRS数据/SECTOREQM/临时查询_查询结果_20250117-新网.xlsx'
    mml_path='G:/工作内容/PRS数据/SECTOREQM/MML报文解析20250117165124-泰山.xlsx'
    mml_path2='G:/工作内容/PRS数据/SECTOREQM/MML报文解析20250117165425-新网.xlsx'
    temp_select_gc='G:/工作内容/25年每月例行工作/工参/苏州华为4G工参20250417.csv'
    rru_filepath='G:/工作内容/PRS数据/SECTOREQM/工具/副本4G-RRU型号20240113.xlsx'
    write_filepath='G:/工作内容/PRS数据/SECTOREQM/工具/4G-RRU疑似被封堵梳理清单.xlsx'



    read_gc(temp_select_gc)
    read_rru_id(rru_filepath)
    read_mml(mml_path)
    read_temp_selectfile(temp_select_filepath)
    read_mml(mml_path2)
    read_temp_selectfile(temp_select_filepath2)

    for li in TXS:
        append_data_to_excel(write_filepath, '第七批次',  TXS[li])

    # 应用样式所有行
    apply_styles_to_last_two_rows(write_filepath, 2, '第七批次')