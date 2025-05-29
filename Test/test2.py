from openpyxl import load_workbook


def insert_combined_column(file_path, sheet_name):
    # 加载工作簿
    wb = load_workbook(filename=file_path)

    # 获取指定工作表
    sheet = wb[sheet_name]

    # 在A列插入新列（原所有列向右移动）
    sheet.insert_cols(1)

    # 从第1行开始遍历所有行
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1), start=1):
        # 获取原第一列（现在是B列）和第三列（现在是D列）的值
        original_first = str(sheet.cell(row=row_idx, column=2).value) if sheet.cell(row=row_idx,
                                                                                    column=2).value is not None else ""
        original_third = str(sheet.cell(row=row_idx, column=4).value) if sheet.cell(row=row_idx,
                                                                                    column=4).value is not None else ""

        # 组合值
        combined_value = f"{original_first}_{original_third}"

        # 在新A列写入组合值
        sheet.cell(row=row_idx, column=1, value=combined_value)

    # 保存修改
    wb.save(file_path)
    print(f"已在工作表'{sheet_name}'的A列插入组合数据，原始数据保留并右移，文件已保存。")

# 使用示例 - 替换为你的实际文件路径和工作表名
insert_combined_column("C:/Users/24253/Desktop/新建文件夹/弹窗时选择此文件夹/用户通报-实时监控-2025年.xlsx", "DSP用户数")