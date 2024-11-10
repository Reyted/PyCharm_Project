import openpyxl
from openpyxl import Workbook
import os
from pathlib import Path


def create_and_copy_excel():
    # 创建新的工作簿
    new_wb = Workbook()

    # 创建3个sheet
    sheet1 = new_wb.active
    sheet1.title = 'Sheet1'
    new_wb.create_sheet('Sheet2')
    new_wb.create_sheet('Sheet3')

    # 在Sheet1中写入表头
    headers = ['序号', '区县', '基站名', '小区名']  # 修正了表头名称
    for col, header in enumerate(headers, 1):
        sheet1.cell(row=1, column=col, value=header)

    # 获取桌面路径
    desktop = str(Path.home() / "Desktop")

    # 打开源文件（aa.xlsx）
    try:
        source_path = os.path.join(desktop, 'aa.xlsx')
        source_wb = openpyxl.load_workbook(source_path)

        # 遍历源文件中的所有sheet
        row_counter = 2  # 从第二行开始写入数据
        for source_sheet in source_wb.worksheets:
            print(f"正在处理sheet: {source_sheet.title}")

            # 获取当前sheet的表头
            source_headers = [str(cell.value).strip() if cell.value else '' for cell in source_sheet[1]]

            # 创建表头映射
            header_mapping = {}
            for i, header in enumerate(headers):
                try:
                    # 使用更灵活的匹配方式
                    for idx, src_header in enumerate(source_headers):
                        # 检查完全匹配或部分匹配
                        if (header == src_header or
                                (header in src_header and '名' in src_header) or
                                (src_header in header and '名' in header)):
                            header_mapping[i] = idx
                            break
                except ValueError:
                    continue

            # 如果找到匹配的表头，则复制数据
            if header_mapping:
                # 复制数据
                for row in source_sheet.iter_rows(min_row=2):
                    has_data = False
                    current_row_data = {}

                    # 收集当前行的所有数据
                    for new_col, source_col in header_mapping.items():
                        value = row[source_col].value
                        if value is not None:  # 只处理非空值
                            current_row_data[new_col] = value
                            has_data = True

                    # 如果行有数据，则写入
                    if has_data:
                        for new_col, value in current_row_data.items():
                            sheet1.cell(row=row_counter, column=new_col + 1, value=value)
                        row_counter += 1

        # 保存新文件到桌面
        new_file_path = os.path.join(desktop, 'new_excel.xlsx')
        new_wb.save(new_file_path)
        print(f"文件已成功保存到: {new_file_path}")

    except FileNotFoundError:
        print("错误：在桌面上找不到'aa.xlsx'文件")
    except Exception as e:
        print(f"发生错误: {str(e)}")
    finally:
        if 'source_wb' in locals():
            source_wb.close()
        new_wb.close()


if __name__ == "__main__":
    create_and_copy_excel()