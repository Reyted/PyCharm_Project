import os
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook


def merge_large_excel_by_sheet(input_folder, output_file, chunksize=5000):
    """
    解决隐藏工作表问题的大文件合并工具

    参数:
        input_folder: 输入文件夹路径
        output_file: 输出文件路径
        chunksize: 分块大小(行)
    """
    excel_files = [f for f in os.listdir(input_folder) if f.endswith(('.xlsx', '.xls'))]

    if not excel_files:
        print("没有找到Excel文件")
        return

    # 第一次遍历：收集所有可见工作表名称
    print("正在扫描工作表结构...")
    visible_sheets = set()
    for file in tqdm(excel_files):
        file_path = os.path.join(input_folder, file)
        try:
            wb = load_workbook(file_path)
            for sheet in wb:
                if not sheet.sheet_state == 'hidden':
                    visible_sheets.add(sheet.title)
            wb.close()
        except Exception as e:
            print(f"扫描文件 {file} 时出错: {str(e)}")

    if not visible_sheets:
        print("错误: 所有工作表都是隐藏的！")
        return

    print(f"找到 {len(visible_sheets)} 个可见工作表")

    # 创建输出文件时确保至少一个工作表可见
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 先创建一个临时可见工作表
        pd.DataFrame().to_excel(writer, sheet_name="临时表", index=False)

        for sheet_name in visible_sheets:
            print(f"\n正在合并: {sheet_name}")
            first_chunk = True

            for file in tqdm(excel_files):
                file_path = os.path.join(input_folder, file)

                try:
                    # 检查该文件是否有此可见工作表
                    wb = load_workbook(file_path)
                    if sheet_name not in wb.sheetnames:
                        wb.close()
                        continue

                    sheet = wb[sheet_name]
                    if sheet.sheet_state == 'hidden':
                        wb.close()
                        continue
                    wb.close()

                    # 分块读取
                    reader = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        chunksize=chunksize
                    )

                    for chunk in reader:
                        chunk['来源文件'] = os.path.basename(file_path)

                        if first_chunk:
                            # 写入新工作表
                            chunk.to_excel(
                                writer,
                                sheet_name=sheet_name,
                                index=False
                            )
                            first_chunk = False
                        else:
                            # 追加到现有工作表
                            book = load_workbook(output_file)
                            writer.book = book
                            writer.sheets = {ws.title: ws for ws in book.worksheets}

                            start_row = book[sheet_name].max_row
                            chunk.to_excel(
                                writer,
                                sheet_name=sheet_name,
                                startrow=start_row,
                                header=False,
                                index=False
                            )
                            writer.save()

                except Exception as e:
                    print(f"处理 {file} 时出错: {str(e)}")
                    continue

        # 删除临时工作表
        book = load_workbook(output_file)
        if "临时表" in book.sheetnames:
            del book["临时表"]
        book.save(output_file)
        book.close()

    print(f"\n合并完成！文件已保存到: {output_file}")


if __name__ == "__main__":
    input_folder = "C:/Users/24253/Desktop/新建文件夹 (9)/新建文件夹"  # 替换为你的输入文件夹路径
    output_file = "merged_output.xlsx"  # 替换为你想要的输出文件名
    merge_large_excel_by_sheet(input_folder, output_file)
