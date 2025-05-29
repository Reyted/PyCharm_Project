import csv
import pandas as pd
import openpyxl
import chardet
import os
import tkinter as tk
from tkinter import Tk, filedialog
from openpyxl import load_workbook
import csv
from copy import copy

def list_files_recursively(path):
    """递归列出路径下的所有文件"""
    file_list = []
    for root, dirs, files in os.walk(path):
        for file in files:
            file_path = os.path.join(root, file)
            file_list.append(file_path)
    return file_list

def select_directory():
    """使用Tkinter选择目录"""
    root = Tk()
    root.withdraw()  # 隐藏主窗口
    folder_path = filedialog.askdirectory(title="请选择要遍历的文件夹")
    root.destroy()
    return folder_path

def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
    return result['encoding']

def list_files_recursively(path):
    """递归列出路径下的所有文件"""
    file_list = []
    for root, dirs, files in os.walk(path):
        for file in files:
            file_path = os.path.join(root, file)
            file_list.append(file_path)
    return file_list

def write_LTE(source_file_path, target_file_path):
    wb = openpyxl.load_workbook(source_file_path)
    TJ = wb['天级']
    XSJ = wb['小时级']
    all_files=list_files_recursively(target_file_path)
    for file in all_files:
        if 'NR' not in file:
            if '小时' in file:
                if '泰山' in file:
                    XSJ_path_dic['泰山']=file
                elif '新网' in file:
                    XSJ_path_dic['新网']=file
                else:
                    XSJ_path_dic['高铁']=file

            if '天级' in file:
                if '泰山' in file:
                    TJ_path_dic['泰山'] = file
                elif '新网' in file:
                    TJ_path_dic['新网'] = file
                else:
                    TJ_path_dic['高铁'] = file
    # print(XSJ_path_dic)
    # print(TJ_path_dic)

    try:
        TS_TJ = openpyxl.load_workbook(TJ_path_dic['泰山'])
        TS_TJ_LTE = TS_TJ['天级']
        for li in TS_TJ_LTE.values:
            # print(li)
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, '泰山')
            if temp_lst[0] != '日期':
                TJ.append(temp_lst)
    except:
        pass

    try:
        XW_TJ = openpyxl.load_workbook(TJ_path_dic['新网'])
        XW_TJ_LTE = XW_TJ['子报表 1']
        for li in XW_TJ_LTE.values:
            # print(li)
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, '新网')
            if temp_lst[0] != '日期':
                TJ.append(temp_lst)
    except:
        pass

    try:
        GT_TJ = openpyxl.load_workbook(TJ_path_dic['高铁'])
        GT_TJ_LTE = GT_TJ['子报表 1']
        for li in GT_TJ_LTE.values:
            # print(li)
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, '高铁')
            if temp_lst[0] != '日期':
                TJ.append(temp_lst)
    except:
        pass
    try:
        TS_XSJ = openpyxl.load_workbook(XSJ_path_dic['泰山'])
        TS_XSJ_LTE = TS_XSJ['子报表 1']
        for li in TS_XSJ_LTE.values:
            #print(li)
            temp_lst=list(li)[3:]
            temp_lst.insert(0,li[0])
            temp_lst.insert(1,'泰山')
            if temp_lst[0] !='时间':
                XSJ.append(temp_lst)
    except:
        pass
    try:
        XW_XSJ = openpyxl.load_workbook(XSJ_path_dic['新网'])
        XW_XSJ_LTE = XW_XSJ['子报表 1']
        for li in XW_XSJ_LTE.values:
            # print(li)
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, '新网')
            if temp_lst[0] != '时间':
                XSJ.append(temp_lst)
    except:
        pass
    try:
        GT_XSJ = openpyxl.load_workbook(XSJ_path_dic['高铁'])
        GT_XSJ_LTE = GT_XSJ['子报表 1']
        for li in GT_XSJ_LTE.values:
            # print(li)
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, '高铁')
            if temp_lst[0] != '时间':
                XSJ.append(temp_lst)
    except:
        pass

    wb.save(source_file_path)

def write_NR(source_file_path, target_file_path):
    wb = openpyxl.load_workbook(source_file_path)
    SA_TJ = wb['SA天级']
    NSA_TJ = wb['NSA天级']
    SA_XSJ = wb['SA小时级']
    NSA_XSJ = wb['NSA小时级']
    all_files = list_files_recursively(target_file_path)
    for file in all_files:
        #print(file)
        if '天级' in file:
            #print(file)
            if "SA" in file and '新网' in file:
                SA_TJ_path_dic['XWSA天级']=file
            if "SA" in file and '泰山' in file:
                SA_TJ_path_dic['TSSA天级']=file
            if "NSA" in file and '新网' in file:
                NSA_TJ_path_dic['XWNSA天级']=file
            if "NSA" in file and '泰山' in file:
                NSA_TJ_path_dic['TSNSA天级']=file
        if '小时级' in file:
            if "SA" in file and '新网' in file:
                SA_XSJ_path_dic['XWSA小时级']=file
            if "SA" in file and '泰山' in file:
                SA_XSJ_path_dic['TSSA小时级']=file
            if "NSA" in file and '新网' in file:
                NSA_XSJ_path_dic['XWNSA小时级']=file
            if "NSA" in file and '泰山' in file:
                NSA_XSJ_path_dic['TSNSA小时级']=file



    try:
        TS_NSA_TJ = openpyxl.load_workbook(NSA_TJ_path_dic['TSNSA天级'])
        TS_TJ_NSA = TS_NSA_TJ['子报表 1']
        for li in TS_TJ_NSA.values:
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, li[1])
            temp_lst.insert(2, '泰山_NSA场景')
            if temp_lst[0] != '日期':
                NSA_TJ.append(temp_lst)
    except:
        pass
    try:
        XW_NSA_TJ = openpyxl.load_workbook(NSA_TJ_path_dic['XWNSA天级'])
        XW_TJ_NSA = XW_NSA_TJ['子报表 1']
        for li in XW_TJ_NSA.values:
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, li[1])
            temp_lst.insert(2, '新网_NSA场景')
            if temp_lst[0] != '日期':
                NSA_TJ.append(temp_lst)
    except:
        pass

    try:
        TS_SA_TJ = openpyxl.load_workbook(SA_TJ_path_dic['TSSA天级'])
        TS_TJ_SA = TS_SA_TJ['子报表 1']
        for li in TS_TJ_SA.values:
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, li[1])
            temp_lst.insert(2, '泰山场景')
            if temp_lst[0] != '日期':
                SA_TJ.append(temp_lst)
    except:
        pass

    try:
        XW_SA_TJ = openpyxl.load_workbook(SA_TJ_path_dic['XWSA天级'])
        XW_TJ_SA = XW_SA_TJ['子报表 1']
        for li in XW_TJ_SA.values:
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, li[1])
            temp_lst.insert(2, '新网场景')
            if temp_lst[0] != '日期':
                SA_TJ.append(temp_lst)
    except:
        pass
    try:
        XW_SA_XSJ = openpyxl.load_workbook(SA_XSJ_path_dic['XWSA小时级'])
        XW_XSJ_SA = XW_SA_XSJ['子报表 1']
        for li in XW_XSJ_SA.values:
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, li[1])
            temp_lst.insert(2, '新网场景')
            if temp_lst[0] != '时间':
                SA_XSJ.append(temp_lst)
    except:
        pass
    try:
        TS_SA_XSJ = openpyxl.load_workbook(SA_XSJ_path_dic['TSSA小时级'])
        TS_XSJ_SA = TS_SA_XSJ['子报表 1']
        for li in TS_XSJ_SA.values:
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, li[1])
            temp_lst.insert(2, '泰山场景')
            if temp_lst[0] != '时间':
                SA_XSJ.append(temp_lst)
    except:
        pass
    try:
        XW_NSA_XSJ = openpyxl.load_workbook(NSA_XSJ_path_dic['XWNSA小时级'])
        XW_XSJ_NSA = XW_NSA_XSJ['子报表 1']
        for li in XW_XSJ_NSA.values:
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, li[1])
            temp_lst.insert(2, '新网_NSA场景')
            if temp_lst[0] != '时间':
                NSA_XSJ.append(temp_lst)
    except:
        pass
    try:
        TS_NSA_XSJ = openpyxl.load_workbook(NSA_XSJ_path_dic['TSNSA小时级'])
        TS_XSJ_NSA = TS_NSA_XSJ['子报表 1']
        for li in TS_XSJ_NSA.values:
            temp_lst = list(li)[3:]
            temp_lst.insert(0, li[0])
            temp_lst.insert(1, li[1])
            temp_lst.insert(2, '泰山_NSA场景')
            if temp_lst[0] != '时间':
                NSA_XSJ.append(temp_lst)
    except:
        pass
    wb.save(source_file_path)

def get_second_row_styles(excel_file,num, sheet_name='Sheet1'):
    # 加载 Excel 文件
    workbook = load_workbook(excel_file)
    # 选择指定的工作表
    sheet = workbook[sheet_name]
    second_row_styles = []
    # 遍历第二行的每个单元格
    for cell in sheet[3]:
        # 复制 StyleProxy 对象
        font = copy(cell.font)
        border = copy(cell.border)
        fill = copy(cell.fill)
        number_format = copy(cell.number_format)
        alignment = copy(cell.alignment)
        # 存储复制后的 StyleProxy 对象的元组
        style = (font, border, fill, number_format, alignment)
        second_row_styles.append(style)
    return second_row_styles

def apply_styles_to_last_two_rows(excel_file, num, sheet_name='Sheet1'):
    # 获取第二行的样式
    second_row_styles = get_second_row_styles(excel_file, num, sheet_name)
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]
    # 获取最后两行的行号
    max_row = sheet.max_row
    last_two_rows = [max_row - 1, max_row]
    # 应用样式到最后两行
    for row_num in last_two_rows:
        for i, style in enumerate(second_row_styles):
            new_cell = sheet.cell(row=row_num, column=i + 1)
            new_cell.font, new_cell.border, new_cell.fill, new_cell.number_format, new_cell.alignment = style
    # 保存修改后的 Excel 文件
    workbook.save(excel_file)




if __name__ == '__main__':
    source_file=''
    source_file2=''
    # target_file='C:/Users/24253/Desktop/新建文件夹 (2)'
    # target_file2='C:/Users/24253/Desktop/新建文件夹 (2)'
    XSJ_path_dic={}
    TJ_path_dic={}

    folder_path = select_directory()
    target_file=folder_path
    target_file2=folder_path
    lst = list_files_recursively(folder_path)
    for file in lst:
        #print(file)
        if 'NR早晚指标监控' in file:
            source_file2=file
        if 'LTE早晚指标' in file and '泰山' not in file and '新网' not in file:
            source_file=file

    SA_XSJ_path_dic = {}
    NSA_XSJ_path_dic = {}
    SA_TJ_path_dic = {}
    NSA_TJ_path_dic = {}

    write_NR(source_file2, target_file2)

    write_LTE(source_file,target_file)
    # apply_styles_to_last_two_rows(source_file, 2, sheet_name='天级')
    # apply_styles_to_last_two_rows(source_file2, 2, sheet_name='SA天级')
