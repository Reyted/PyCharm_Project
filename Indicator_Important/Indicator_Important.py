import os
import pandas as pd
import openpyxl
import csv
from copy import copy
from openpyxl import load_workbook
from datetime import datetime

all_list=[]
count_list=[]
sheet2_list=[]
sheet2_list_data=[]
gc_band=dict()
band_count={'n41':0,'n28':0}
band_count_sinr={'n41':0,'n28':0}
gc_path='G:/工作内容/25年每月例行工作/工参/苏州华为工参20241223/苏州华为5G工参20241223.csv'
file_path_prs='G:/工作内容/PRS数据'
file_path_xlsx='G:/工作内容/25年每月例行工作/1月工作内容/重点指标数据/2月/06/指标统计.xlsx'


def list_files(directory):
    file_list = []
    # 列出指定目录下的文件和目录
    items = os.listdir(directory)
    for item in items:
        # 构建完整的文件或目录路径
        full_path = os.path.join(directory, item)
        if os.path.isfile(full_path):
            # 如果是文件，添加到文件列表中
            file_list.append(full_path)
    for item in file_list:
        read_excel(item)
    sheet2_list_data=list(sheet2_list[0])+list(sheet2_list[1])
    #print(len(sheet2_list_data))

#读工参
def read_csv_file(file_path):
    data = []
    try:
        with open(file_path, mode='r') as file:
            csv_reader = csv.reader(file)
            for row in csv_reader:
                #data.append(row)
                #print(row)
                gc_band[row[8]] = row[19]

        # for li in data:
        #     #gc_band.append({'gnodeb_name':li[8],'band':li[19]})

        #print(len(gc_band))
    except FileNotFoundError:
        print(f"文件 {file_path} 未找到，请检查文件路径是否正确。")
    except Exception as e:
        print(f"读取文件时发生错误: {e}")
    #print(len(data))

#read_csv_file(gc_path)


def read_excel(file_path):
    df=pd.read_excel(file_path,sheet_name='子报表 1')
    df2=pd.read_excel(file_path,sheet_name='子报表 2')

    read_csv_file(gc_path)

    for li in df2.values:
        #计算频段个数
        lst_li=list(li)
        try:
            if gc_band[lst_li[1]] == 'n41':
                band_count['n41'] += 1
            if gc_band[lst_li[1]] == 'n28':
                band_count['n28'] += 1
        except:
            pass
        #计算各频段干扰个数
        try:
            sinr_num=float(lst_li[6])
            if gc_band[lst_li[1]] == 'n41' and sinr_num > -107.0:
                band_count_sinr['n41'] += 1
            if gc_band[lst_li[1]] == 'n28' and sinr_num > -107.0:
                band_count_sinr['n28'] += 1
        except:
            pass
    #print(band_count_sinr)

    sheet2_list.append(list(df2.values))

    write_newdata(df.values)

def write_newdata(data_list:list):
    new_list=[]
    for data in data_list:
        str1=str(data[1]).split('-')
        if str1[1]=='2.6G':
            temp_list=[2.6]+list(data)[3:]
            new_list.append(temp_list)
        else:
            temp_list = [700] + list(data)[3:]
            new_list.append(temp_list)
    all_list.append(new_list)

def count_data():
    two_list=[]
    seven_list=[]
    #700M
    for i in range(1,4):
        seven_list.append((all_list[0][1][i]+all_list[1][1][i])/2)
    for i in range(4,7):
        seven_list.append(all_list[0][1][i]+all_list[1][1][i])
    for i in range(7,len(all_list[0][1])):
        seven_list.append((all_list[0][1][i]+all_list[1][1][i])/2)
    #2.6
    for i in range(1,4):
        two_list.append((all_list[0][0][i]+all_list[1][0][i])/2)
    for i in range(4,7):
        two_list.append(all_list[0][0][i]+all_list[1][0][i])
    for i in range(7,len(all_list[0][1])):
        two_list.append((all_list[0][0][i]+all_list[1][0][i])/2)

    #print(band_count_sinr['n28']/band_count['n28'])
    time_str=get_time()


    count_list.append([time_str,'华为',700,band_count['n28']]+seven_list[:12]+[band_count_sinr['n28']/band_count['n28']]+seven_list[12:])
    count_list.append([time_str,'华为',2.6,band_count['n41']]+two_list[:12]+[band_count_sinr['n41']/band_count['n41']]+two_list[12:])
    #print(two_list[12:])
    #print(count_list)


def append_data_to_excel(file_path, sheet_name,sheet_name2, data,data2):
    # 加载现有的 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表
    sheet = workbook[sheet_name]
    sheet2 = workbook[sheet_name2]
    # 获取工作表中现有的最大行数
    row_index = sheet.max_row
    row_index2 = sheet2.max_row
    # 遍历要追加的数据列表
    for row in data:
        # 将数据逐行添加到工作表的下一行
        sheet.append(row)
    for row in data2:
        # 将数据逐行添加到工作表的下一行
        sheet2.append(row)
    # 保存更新后的 Excel 文件
    workbook.save(file_path)

def get_time():
    # 原始时间字符串
    original_time_str = sheet2_list[0][0][0]
    # 将字符串解析为 datetime 对象
    datetime_obj = datetime.strptime(original_time_str, '%Y-%m-%d')
    # 获取年
    year = datetime_obj.year
    # 获取月
    month = datetime_obj.month
    # 获取日
    day = datetime_obj.day
    return str(year)+'/'+str(month)+'/'+str(day)


#添加样式
def get_second_row_styles(excel_file,num, sheet_name='Sheet1'):
    # 加载 Excel 文件
    workbook = load_workbook(excel_file)
    # 选择指定的工作表
    sheet = workbook[sheet_name]
    second_row_styles = []
    # 遍历第二行的每个单元格
    for cell in sheet[3]:
        # 复制 StyleProxy 对象
        font = copy(cell.font)
        border = copy(cell.border)
        fill = copy(cell.fill)
        number_format = copy(cell.number_format)
        alignment = copy(cell.alignment)
        # 存储复制后的 StyleProxy 对象的元组
        style = (font, border, fill, number_format, alignment)
        second_row_styles.append(style)
    return second_row_styles


def apply_styles_to_last_two_rows(excel_file,num, sheet_name='Sheet1'):
    # 获取第二行的样式
    second_row_styles = get_second_row_styles(excel_file,num, sheet_name)
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]
    # 获取最后两行的行号
    max_row = sheet.max_row
    last_two_rows = [max_row - 1, max_row]
    # 应用样式到最后两行
    for row_num in last_two_rows:
        for i, style in enumerate(second_row_styles):
            new_cell = sheet.cell(row=row_num, column=i + 1)
            new_cell.font, new_cell.border, new_cell.fill, new_cell.number_format, new_cell.alignment = style
    # 保存修改后的 Excel 文件
    workbook.save(excel_file)



if __name__ == "__main__":
    # 调用 list_files 函数并传入要读取的目录，例如当前目录 '.'
    list_files(file_path_prs)# 要操作的工作表名称
    sheet_name = 'Sheet1'
    sheet_name2 = 'Sheet2'
    count_data()


    #sheet1
    data_list=[count_list[0][:14]+[count_list[0][16]],
               count_list[1][:14]+[count_list[1][16]]]
    data_list2=[count_list[0],count_list[1]]
    # print(data_list2)
    # print(count_list)
    append_data_to_excel(file_path_xlsx, sheet_name,sheet_name2, data_list,data_list2)
    # #append_data_to_excel(file_path_xlsx, sheet_name2, data_list2)
    #
    # # 应用样式到最后两行
    apply_styles_to_last_two_rows(file_path_xlsx,4656,sheet_name)
    apply_styles_to_last_two_rows(file_path_xlsx,2260,sheet_name2)