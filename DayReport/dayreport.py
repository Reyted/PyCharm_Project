import openpyxl
from openpyxl import load_workbook
import pandas as pd
import datetime

def convert_number_to_letter(num):
    return chr(ord('A') + num-1)

def write_report(filepath:str,ind1:int):
    wb = openpyxl.load_workbook(filepath)
    sheet1 = wb.create_sheet('每日汇报1')
    sheet = wb['每日汇总']

    df1 = pd.read_excel(filepath, sheet_name='10月')
    df2 = pd.read_excel(filepath, sheet_name='每日汇总')

    today_list = []
    townlist = ['卡若区', '八宿县', '边坝县', '察雅县', '丁青县', '贡觉县', '江达县', '类乌齐县', '洛隆县', '芒康县',
                '左贡县']

    base_station=[]
    conplaint=dict()

    #datestr = str((datetime.datetime.now()).strftime('%Y-%m-%d'))
    datestr = '2024-10-19'
    town_obj = {'卡若区': 0, '八宿县': 0, '边坝县': 0, '察雅县': 0, '丁青县': 0, '贡觉县': 0, '江达县': 0,
                '类乌齐县': 0, '洛隆县': 0, '芒康县': 0, '左贡县': 0}
    town_obj_month = {'卡若区': 0, '八宿县': 0, '边坝县': 0, '察雅县': 0, '丁青县': 0, '贡觉县': 0, '江达县': 0,
                      '类乌齐县': 0, '洛隆县': 0, '芒康县': 0, '左贡县': 0}
    solve_obj = {'卡若区': 0, '八宿县': 0, '边坝县': 0, '察雅县': 0, '丁青县': 0, '贡觉县': 0, '江达县': 0,
                 '类乌齐县': 0, '洛隆县': 0, '芒康县': 0, '左贡县': 0}
    town_obj_top_list = {'卡若区': [], '八宿县': [], '边坝县': [], '察雅县': [], '丁青县': [], '贡觉县': [],
                         '江达县': [], '类乌齐县': [], '洛隆县': [], '芒康县': [], '左贡县': []}
    town_obj_top = {'卡若区': '', '八宿县': '', '边坝县': '', '察雅县': '', '丁青县': '', '贡觉县': '', '江达县': '',
                    '类乌齐县': '', '洛隆县': '', '芒康县': '', '左贡县': ''}
    unsolve_obj_list = dict()
    base_station_count = dict()
    issolve_basestation=dict()
    issolve_basestation_n=dict()
    issolve_basestation_y=dict()
    sort_base_station_count_n=dict()
    sort_base_station_count_y=dict()

    for row in df1.values:
        dt = str(row[0]).split(' ')[0]
        if dt == datestr:
            issolve_basestation[row[11]]=row[17]
            if row[17]=='解决' :
                issolve_basestation_y[row[11]] = row[17]
            else:
                issolve_basestation_n[row[11]] = row[17]

            conplaint[row[11]]=row[12]
            base_station.append(row[11])
            today_list.append(row)
            town_obj_top_list[row[8]].append(row[11].split('_')[2])

    for i in base_station:
        if i not in base_station_count:
            base_station_count[i]=1
        else:
            base_station_count[i]+=1

    for row in df1.values:
        town_obj_month[row[8]] += 1

    for li in today_list:
        for town in townlist:
            if li[8] == town:
                town_obj[town] += 1
                if li[17] == '解决':
                    solve_obj[town] += 1

    for town in town_obj_top_list:
        nowstr = ''
        nexrstr = ''
        ind = 0
        m = {}
        tp = tuple()
        for li in town_obj_top_list[town]:
            if li in m:
                m[li] += 1
            else:
                m[li] = 0
        if len(m) > 0:
            tp = sorted(m.items(), key=lambda x: x[1])[len(sorted(m.items(), key=lambda x: x[1])) - 1]
        if len(tp) > 0:
            town_obj_top[town] = tp[0]
        else:
            town_obj_top[town] = ''

    sort_town_obj = dict(sorted(town_obj.items(), key=lambda item: item[1]))
    sort_town=dict()

    for i in town_obj:
        for j in solve_obj:
            unsolve_obj_list[i]=town_obj[i]-solve_obj[i]

    indd=11
    for i in sort_town_obj:
        sort_town[i]=indd
        indd-=1

    l=['A','B','C','D','E','F','G']
    s=[1,2,3,4,5,6]
    index=0
    all_data = [town_obj, solve_obj, unsolve_obj_list,sort_town, town_obj_month, town_obj_top_list]
    for row in df2.values:
        if index >= ind1:
            town = row[0]
            if town!='全量':
                for i in s:
                    letter=convert_number_to_letter(i+1)
                    try:
                        sheet[f'{letter}{index + 2}'] = '、'.join(set(all_data[i - 1][town]))
                    except:
                        sheet[f'{letter}{index + 2}'] = all_data[i - 1][town]
        index += 1

    sheet1.append(['基站名','投诉信息','投诉频次','当前是否解决','备注','日期'])
    new_base_station_count=sorted(base_station_count.items(), key=lambda item: item[1],reverse = True)

    # 假设你有一个包含字典的列表
    list_of_dicts = [
        issolve_basestation_y,
        issolve_basestation_n
    ]

    # 使用列表推导式和字典的转换功能去除重复项
    unique_dicts = list({tuple(sorted(d.items())) for d in list_of_dicts})

    # 将元组转回字典
    unique_dicts = [dict(t) for t in unique_dicts]

    new_base_station_count_dict=dict(new_base_station_count)


    for item in unique_dicts[0]:
        sort_base_station_count_n[item]=new_base_station_count_dict[item]
    for item in unique_dicts[1]:
        sort_base_station_count_y[item]=new_base_station_count_dict[item]


    new_sort_base_station_count_y=sorted(sort_base_station_count_y.items(), key=lambda item: item[1], reverse=True)
    new_sort_base_station_count_n=sorted(sort_base_station_count_n.items(), key=lambda item: item[1], reverse=True)


    for basestation in new_sort_base_station_count_y:
        try:
            sheet1.append(
                [basestation[0], conplaint[basestation[0]], new_base_station_count_dict[basestation[0]], issolve_basestation[basestation[0]], '',
                 datestr])
        except:
            pass
    for basestation in new_sort_base_station_count_n:
        try:
            sheet1.append(
                [basestation[0], conplaint[basestation[0]], new_base_station_count_dict[basestation[0]],
                 issolve_basestation[basestation[0]], '',
                 datestr])
        except:
            pass


    wb.save(filepath)


if __name__ == '__main__':
    write_report('C:/Users/24253/Desktop/Python_Excel/副本昌都故障类投诉汇总20241014(2).xlsx',46)








