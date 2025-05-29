import pandas as pd
import csv
import os
from openpyxl import load_workbook
from datetime import datetime
import openpyxl
from copy import copy
import threading
import time
import os
import chardet
# day='16'
# jieguo_url='G:/工作内容/25年每月例行工作/1月工作内容/冗余邻区/3月第一周/'
# update_file_path = f'{jieguo_url}删除冗余核查05{day}.xlsx'

HW5GGC_PQ=dict()
HW5GGC_WG=dict()
HW4GGC_PQ=dict()
HW4GGC_WG=dict()
HW4GGC_index=dict()
CELL_STATIC=[]
ENODEB_ID=dict()

COUNT=[{},{},{},{}]

# workbook = openpyxl.load_workbook(update_file_path)
# # 选择要操作的工作表
# sheet2 = workbook['5-5冗余']
# sheet3 = workbook['5-4冗余']
# sheet4 = workbook['4-4冗余']
# sheet5 = workbook['4-5冗余']

def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
    return result['encoding']

def read_HW5GGC_file(file_path):
    #encoding = detect_encoding(file_path)
    with open(file_path, 'r', encoding='gbk') as f:
        reader = csv.reader(f)
        for row in reader:
            try:
                if HW5GGC_PQ[row[8]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                    pass
            except:
                HW5GGC_PQ[row[8]] = row[5]
                HW5GGC_WG[row[8]] = row[39]

def read_HW4GGC_file(file_path):
    #encoding = detect_encoding(file_path)
    global HW4GGC_index
    with open(file_path, 'r',encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            try:
                if HW4GGC_PQ[row[1]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                    pass
            except:
                HW4GGC_PQ[row[1]] = row[4]
                HW4GGC_WG[row[1]] = row[39]
            try:
                HW4GGC_index[str(row[21]) + '_' + str(row[17])] += 1
            except:
                HW4GGC_index[str(row[21]) + '_' + str(row[17])] = 1

def read_cellstatic_file(file_path1, file_path2):
    with open(file_path2, 'r', encoding='gbk') as f:
        reader = csv.reader(f)
        for row in reader:
            ENODEB_ID[row[1]]=row[6]
    with open(file_path1, 'r', encoding='gbk') as f:
        reader = csv.reader(f)
        for row in reader:
            try:
                CELL_STATIC.append(ENODEB_ID[row[1]] + "-" + row[17])
            except:
                pass

sheet2_lst=[]
def read_fivetofivery_file(file_path):
    pf=pd.read_excel(file_path)
    ind = len(sheet2_lst)
    for row in pf.values:
        try:
            if HW5GGC_PQ[row[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                #temp_lst=[]
                try:
                    COUNT[0][HW5GGC_PQ[row[0]]] += 1
                except:
                    COUNT[0][HW5GGC_PQ[row[0]]] = 1
                sheet2_lst.append(list(row[:7]))
                #temp_lst=list(row[:7])
                sheet2_lst[ind].append('')
                #temp_lst.append('')
                sheet2_lst[ind].append('')
                #temp_lst[ind].append('')
                sheet2_lst[ind].append('')
                #temp_lst[ind].append('')
                if HW5GGC_WG[row[0]] in '2288X泰山' and HW5GGC_WG[row[0]] !='':
                    sheet2_lst[ind].append(HW5GGC_PQ[row[0]])
                    #temp_lst[ind].append(HW5GGC_PQ[row[0]])
                    sheet2_lst[ind].append(HW5GGC_WG[row[0]])
                    #temp_lst.append(HW5GGC_WG[row[0]])
                else:
                    if HW5GGC_PQ[row[0]] == '中移3':
                        sheet2_lst[ind].append(HW5GGC_PQ[row[0]])
                        #temp_lst.append(HW5GGC_PQ[row[0]])
                        sheet2_lst[ind].append('泰山')
                        #temp_lst.append('泰山')
                    else:
                        sheet2_lst[ind].append(HW5GGC_PQ[row[0]])
                        #temp_lst.append(HW5GGC_PQ[row[0]])
                        sheet2_lst[ind].append('2288X')
                        #temp_lst.append('2288X')
                ind += 1
                #sheet2.append(temp_lst)
        except:
            pass

sheet3_lst=[]
def read_fivetofourry_file(file_path1,file_path2):
    pf1=pd.read_excel(file_path1)
    pf2=pd.read_excel(file_path2)
    tempory_dic = dict()
    def tempory(li):
        try:
            if HW5GGC_PQ[li[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                try:
                    COUNT[1][HW5GGC_PQ[li[0]]] += 1
                except:
                    COUNT[1][HW5GGC_PQ[li[0]]] = 1
                li.append('是')
                li.append('')
                li.append('')
                li.append('')
                if HW5GGC_WG[li[0]] in '2288X泰山' and HW5GGC_WG[li[0]] !='':
                    li.append(HW5GGC_PQ[li[0]])
                    li.append(HW5GGC_WG[li[0]])
                else:
                    if HW5GGC_PQ[li[0]] == '中移3':
                        li.append(HW5GGC_PQ[li[0]])
                        li.append('泰山')
                    else:
                        li.append(HW5GGC_PQ[li[0]])
                        li.append('2288X')
                #sheet3.append(li)
                sheet3_lst.append(li)
        except:
            pass
    for row in pf2.values:
        if str(row[1]) + "-" + str(row[2]) not in CELL_STATIC:
            st = row[0] + str(row[1]) + str(row[2])
            if st not in tempory_dic:
                tempory_dic[st] = 1
                tempory(list(row[:7]))
        # try:
        #     if HW4GGC_index[str(row[1]) + '_' + str(row[2])] != 2:
        #         if str(row[1])+"-"+str(row[2]) not in CELL_STATIC:
        #             st = row[0] + str(row[1]) + str(row[2])
        #             if st not in tempory_dic:
        #                 tempory_dic[st] = 1
        #                 tempory(list(row[:7]))
        # except:
        #     if str(row[1]) + "-" + str(row[2]) not in CELL_STATIC:
        #         st = row[0] + str(row[1]) + str(row[2])
        #         if st not in tempory_dic:
        #             tempory_dic[st] = 1
        #             tempory(list(row[:7]))
    for row in pf1.values:
        st = row[0] + str(row[1]) + str(row[2])
        if st not in tempory_dic:
            if str(row[1]) + "-" + str(row[2]) not in CELL_STATIC:
                tempory_dic[st] = 1
                tempory(list(row[:7]))
        # try:
        #     if HW4GGC_index[str(row[1]) + '_' + str(row[2])] != 2:
        #         st = row[0] + str(row[1]) + str(row[2])
        #         if st not in tempory_dic:
        #             if str(row[1])+"-"+str(row[2]) not in CELL_STATIC:
        #                 tempory_dic[st] = 1
        #                 tempory(list(row[:7]))
        # except:
        #     st = row[0] + str(row[1]) + str(row[2])
        #     if st not in tempory_dic:
        #         if str(row[1]) + "-" + str(row[2]) not in CELL_STATIC:
        #             tempory_dic[st] = 1
        #             tempory(list(row[:7]))
sheet4_lst=[]
def read_fourtofourry_file(file_path1,file_path2):
    pf1 = pd.read_excel(file_path1)
    pf2 = pd.read_excel(file_path2)
    tempory_dic = dict()
    def tempory(li):
        try:
            if HW4GGC_PQ[li[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                try:
                    COUNT[2][HW4GGC_PQ[li[0]]] += 1
                except:
                    COUNT[2][HW4GGC_PQ[li[0]]] = 1
                li.append('是')
                li.append('')
                li.append('')
                li.append('')
                li.append('')
                if HW4GGC_WG[li[0]] in '2288X泰山' and HW4GGC_WG[li[0]]!='':
                    li.append(HW4GGC_PQ[li[0]])
                    li.append(HW4GGC_WG[li[0]])
                else:
                    if HW4GGC_PQ[li[0]] == '中移3':
                        li.append(HW4GGC_PQ[li[0]])
                        li.append('泰山')
                    else:
                        li.append(HW4GGC_PQ[li[0]])
                        li.append('2288X')
                sheet4_lst.append(li)
                #sheet4.append(li)
        except:
            pass
    for row in pf2.values:
        if str(row[1]) + "-" + str(row[2]) not in CELL_STATIC:
            st = row[0] + str(row[1]) + str(row[2])
            if st not in tempory_dic:
                temp_lst = list(row[:3]) + [row[10]] + list(row[3:6])
                tempory_dic[st] = 1
                tempory(temp_lst)
        # try:
        #     if HW4GGC_index[str(row[1]) + '_' + str(row[2])] != 2:
        #         if str(row[1]) + "-" + str(row[2]) not in CELL_STATIC:
        #             st=row[0]+str(row[1])+str(row[2])
        #             if st not in tempory_dic:
        #                 temp_lst=list(row[:3])+[row[10]]+list(row[3:6])
        #                 tempory_dic[st]=1
        #                 tempory(temp_lst)
        # except:
        #     if str(row[1]) + "-" + str(row[2]) not in CELL_STATIC:
        #         st = row[0] + str(row[1]) + str(row[2])
        #         if st not in tempory_dic:
        #             temp_lst = list(row[:3]) + [row[10]] + list(row[3:6])
        #             tempory_dic[st] = 1
        #             tempory(temp_lst)
    for row in pf1.values:
        if str(row[1]) + "-" + str(row[2]) not in CELL_STATIC:
            st = row[0] + str(row[1]) + str(row[2])
            if st not in tempory_dic:
                tempory_dic[st] = 1
                tempory(row[:7])
        # try:
        #     if HW4GGC_index[str(row[1]) + '_' + str(row[2])] != 2:
        #         if str(row[1]) + "-" + str(row[2]) not in CELL_STATIC:
        #             st = row[0] + str(row[1]) + str(row[2])
        #             if st not in tempory_dic:
        #                 tempory_dic[st] = 1
        #                 tempory(row[:7])
        # except:
        #     if str(row[1]) + "-" + str(row[2]) not in CELL_STATIC:
        #         st = row[0] + str(row[1]) + str(row[2])
        #         if st not in tempory_dic:
        #             tempory_dic[st] = 1
        #             tempory(row[:7])
sheet5_lst=[]
def read_fourtofivery_file(file_path):
    df = pd.read_excel(file_path)
    ind=len(sheet5_lst)
    for row in df.values:
        try:
            if HW4GGC_PQ[row[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                #temp_lst=[]
                try:
                    COUNT[3][HW4GGC_PQ[row[0]]] += 1
                except:
                    COUNT[3][HW4GGC_PQ[row[0]]] = 1
                sheet5_lst.append(list(row[:7]))
                #temp_lst=list(row[:7])
                sheet5_lst[ind].append('')
                #temp_lst.append('')
                sheet5_lst[ind].append('')
                #temp_lst.append('')
                sheet5_lst[ind].append('')
                #temp_lst.append('')
                if HW4GGC_WG[row[0]] in '2288X泰山' and HW4GGC_WG[row[0]]!='':
                    sheet5_lst[ind].append(HW4GGC_PQ[row[0]])
                    #temp_lst.append(HW4GGC_PQ[row[0]])
                    sheet5_lst[ind].append(HW4GGC_WG[row[0]])
                    #temp_lst.append(HW4GGC_WG[row[0]])
                else:
                    if HW4GGC_PQ[row[0]]=='中移3':
                        sheet5_lst[ind].append(HW4GGC_PQ[row[0]])
                        #temp_lst.append(HW4GGC_PQ[row[0]])
                        sheet5_lst[ind].append('泰山')
                        #temp_lst.append('泰山')
                    else:
                        sheet5_lst[ind].append(HW4GGC_PQ[row[0]])
                        #temp_lst.append(HW4GGC_PQ[row[0]])
                        sheet5_lst[ind].append('2288X')
                        #temp_lst.append('2288X')
                ind+=1
                #sheet5.append(temp_lst)
        except:
            pass

#添加数据到excel
def append_data_to_excel(file_path,sheet_name2,sheet_name3,sheet_name4,sheet_name5,data2,data3,data4,data5):
    #加载现有的 Excel 文件
    # workbook = openpyxl.load_workbook(file_path)
    # # 选择要操作的工作表
    # sheet2 = workbook[sheet_name2]
    # sheet3 = workbook[sheet_name3]
    # sheet4 = workbook[sheet_name4]
    # sheet5 = workbook[sheet_name5]

    #遍历要追加的数据列表
    print('5-5')
    for row in data2:
        # 将数据逐行添加到工作表的下一行
        sheet2.append(row)
    print('5-4')
    for row in data3:
        # 将数据逐行添加到工作表的下一行
        sheet3.append(row)
    print('4-4')
    for row in data4:
        # 将数据逐行添加到工作表的下一行
        sheet4.append(row)
    print('4-5')
    for row in data5:
        # 将数据逐行添加到工作表的下一行
        sheet5.append(row)
    # 保存更新后的 Excel 文件
    workbook.save(file_path)

def write_to_excel(file_path, sheet_name, row, column,value):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    # 获取指定工作表
    sheet = workbook[sheet_name]
    sheet.cell(row=row, column=column, value=value)
    # 在指定单元格写入数据
    #sheet.cell(row=row, column=column, value=value)
    # 保存修改后的 Excel 文件
    workbook.save(file_path)

def write_val(count:list,file_path,sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    # 获取指定工作表
    sheet = workbook['数量']
    ind=1
    for row in count:
        for li in range(11):
            cell = sheet.cell(row=li + 2, column=1)
            value = cell.value
            try:
                write_to_excel(file_path, sheet_name, li + 2, ind + 1, row[value])
            except:
                pass
        ind += 1


if __name__ == '__main__':
    day=29
    #day=input('请输入日期')
    jieguo_url='G:/工作内容/25年每月例行工作/1月工作内容/冗余邻区/3月第一周/'

    fivetofivery_file_path = f'{jieguo_url}结果/5-5核查/5-5冗余外部核查1.xlsx'

    fivetofourry_file_path = f'{jieguo_url}结果/5-4核查/5-4冗余外部核查.xlsx'
    fivetofourwkt_file_path = f'{jieguo_url}结果/5-4核查/5-4外部未开通核查.xlsx'

    fourtofourry_file_path = f'{jieguo_url}结果/4-4核查/4-4冗余外部核查.xlsx'
    fourtofourwkt_file_path = f'{jieguo_url}结果/4-4核查/4-4外部未开通核查.xlsx'

    fourtofivery_file_path = f'{jieguo_url}结果/4-5核查/4-5冗余外部核查.xlsx'

    HW5GGC_file_path = f'{jieguo_url}解析文件/苏州华为5G工参202505{day}.csv'
    HW4GGC_file_path = f'{jieguo_url}解析文件/苏州华为4G工参202505{day}.csv'
    cellstatic_file_path = f'{jieguo_url}解析文件/查询小区静态参数.csv'
    enodbfunction_file_path = f'{jieguo_url}解析文件/查询eNodeB功能配置.csv'
    update_file_path = f'{jieguo_url}删除冗余核查05{day}.xlsx'

    workbook = openpyxl.load_workbook(update_file_path)
    # 选择要操作的工作表
    sheet2 = workbook['5-5冗余']
    sheet3 = workbook['5-4冗余']
    sheet4 = workbook['4-4冗余']
    sheet5 = workbook['4-5冗余']
    sheet6=workbook['数量']

    print('正在核查，请稍等.....')
    read_HW5GGC_file(HW5GGC_file_path)
    read_HW4GGC_file(HW4GGC_file_path)
    read_cellstatic_file(cellstatic_file_path,enodbfunction_file_path)
    read_fivetofivery_file(fivetofivery_file_path)
    read_fivetofourry_file(fivetofourry_file_path,fivetofourwkt_file_path)
    read_fourtofourry_file(fourtofourry_file_path,fourtofourwkt_file_path)
    read_fourtofivery_file(fourtofivery_file_path)
    #workbook.save(update_file_path)
    append_data_to_excel(update_file_path,'5-5冗余','5-4冗余','4-4冗余','4-5冗余',sheet2_lst,sheet3_lst,sheet4_lst,sheet5_lst)
    write_val(COUNT, update_file_path, '数量')
    #write_val(COUNT)
