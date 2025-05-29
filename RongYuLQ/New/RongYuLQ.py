import pandas as pd
import csv
import os
from openpyxl import load_workbook
from datetime import datetime
import openpyxl
from copy import copy
import threading
import time
import os
import chardet

HW5GGC_PQ=dict()
HW5GGC_WG=dict()
HW4GGC_PQ=dict()
HW4GGC_WG=dict()
#enodebid+cellid 出现次数
HW4GGC_index=dict()
EXTERNAL=[]
COUNT=[{},{},{},{}]

def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
    return result['encoding']

def read_HW5GGC_file(file_path):
    #encoding = detect_encoding(file_path)
    with open(file_path, 'r', encoding='gbk') as f:
        reader = csv.reader(f)
        for row in reader:
            try:
                if HW5GGC_PQ[row[8]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                    pass
            except:
                HW5GGC_PQ[row[8]] = row[5]
                HW5GGC_WG[row[8]] = row[39]

def read_HW4GGC_file(file_path):
    #encoding = detect_encoding(file_path)
    global HW4GGC_index
    with open(file_path, 'r',encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            try:
                if HW4GGC_PQ[row[1]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                    pass
            except:
                HW4GGC_PQ[row[1]] = row[4]
                HW4GGC_WG[row[1]] = row[39]
            try:
                HW4GGC_index[str(row[21]) + '_' + str(row[17])] += 1
            except:
                HW4GGC_index[str(row[21]) + '_' + str(row[17])] = 1

sheet2_lst=[]
def read_fivetofiveextrenal_file(file_path,file_path2):
    df = pd.read_excel(file_path)
    df2 = pd.read_excel(file_path2)
    ind=len(sheet2_lst)
    for row in df.values:
        try:
            if HW5GGC_PQ[row[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                try:
                    COUNT[0][HW5GGC_PQ[row[0]]] += 1
                except:
                    COUNT[0][HW5GGC_PQ[row[0]]] = 1
                sheet2_lst.append(list(row[:13]))
                sheet2_lst[ind].append('')
                sheet2_lst[ind].append('')
                if HW5GGC_WG[row[0]] in '2288X泰山' and HW5GGC_WG[row[0]]!='':
                    sheet2_lst[ind].append(HW5GGC_PQ[row[0]])
                    sheet2_lst[ind].append(HW5GGC_WG[row[0]])
                else:
                    if HW5GGC_PQ[row[0]] == '中移3':
                        sheet2_lst[ind].append(HW5GGC_PQ[row[0]])
                        sheet2_lst[ind].append('泰山')
                    else:
                        sheet2_lst[ind].append(HW5GGC_PQ[row[0]])
                        sheet2_lst[ind].append('2288X')
                ind += 1
        except:
            pass
    for row in df2.values:
        if row[13]=='否':
            try:
                if HW5GGC_PQ[row[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                    try:
                        COUNT[0][HW5GGC_PQ[row[0]]] += 1
                    except:
                        COUNT[0][HW5GGC_PQ[row[0]]] = 1
                    sheet2_lst.append(list(row[:13]))
                    sheet2_lst[ind].append('')
                    sheet2_lst[ind].append('')
                    if HW5GGC_WG[row[0]] in '2288X泰山' and HW5GGC_WG[row[0]]!='':
                        sheet2_lst[ind].append(HW5GGC_PQ[row[0]])
                        sheet2_lst[ind].append(HW5GGC_WG[row[0]])
                    else:
                        if HW5GGC_PQ[row[0]] == '中移3':
                            sheet2_lst[ind].append(HW5GGC_PQ[row[0]])
                            sheet2_lst[ind].append('泰山')
                        else:
                            sheet2_lst[ind].append(HW5GGC_PQ[row[0]])
                            sheet2_lst[ind].append('2288X')
                    ind += 1
            except:
                pass
sheet3_lst=[]
def read_fivetofourextrenal_file(file_path,s:str):
    df = pd.read_excel(file_path)
    ind=len(sheet3_lst)
    for row in df.values:
        if s=='hw':
            if HW4GGC_index[str(row[1]) + '_' + str(row[2])] != 2:
                try:
                    if HW5GGC_PQ[row[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                        try:
                            COUNT[1][HW5GGC_PQ[row[0]]] += 1
                        except:
                            COUNT[1][HW5GGC_PQ[row[0]]] = 1
                        sheet3_lst.append(list(row[:13]))
                        sheet3_lst[ind].append('')
                        if HW5GGC_WG[row[0]] in '2288X泰山':
                            sheet3_lst[ind].append(HW5GGC_PQ[row[0]])
                            sheet3_lst[ind].append(HW5GGC_WG[row[0]])
                        else:
                            if HW5GGC_PQ[row[0]] == '中移3':
                                sheet3_lst[ind].append(HW5GGC_PQ[row[0]])
                                sheet3_lst[ind].append('泰山')
                            else:
                                sheet3_lst[ind].append(HW5GGC_PQ[row[0]])
                                sheet3_lst[ind].append('2288X')
                        ind += 1
                except:
                    pass
        else:
            if row[13] == '否':
                try:
                    if HW5GGC_PQ[row[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                        try:
                            COUNT[1][HW5GGC_PQ[row[0]]] += 1
                        except:
                            COUNT[1][HW5GGC_PQ[row[0]]] = 1
                        sheet3_lst.append(list(row[:13]))
                        sheet3_lst[ind].append('')
                        if HW5GGC_WG[row[0]] in '2288X泰山' and HW5GGC_WG[row[0]]!='':
                            sheet3_lst[ind].append(HW5GGC_PQ[row[0]])
                            sheet3_lst[ind].append(HW5GGC_WG[row[0]])
                        else:
                            if HW5GGC_PQ[row[0]] == '中移3':
                                sheet3_lst[ind].append(HW5GGC_PQ[row[0]])
                                sheet3_lst[ind].append('泰山')
                            else:
                                sheet3_lst[ind].append(HW5GGC_PQ[row[0]])
                                sheet3_lst[ind].append('2288X')
                        ind += 1
                except:
                    pass
    print(len(sheet3_lst))
sheet4_lst=[]
def read_fourtofourextrenal_file(file_path,s:str):
    df = pd.read_excel(file_path)
    ind=len(sheet4_lst)
    for row in df.values:
        if row[13]=='否':
            if s=='alx':
                try:
                    if HW4GGC_PQ[row[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                        try:
                            COUNT[2][HW4GGC_PQ[row[0]]] += 1
                        except:
                            COUNT[2][HW4GGC_PQ[row[0]]] = 1
                        sheet4_lst.append(list(row[:13]))
                        sheet4_lst[ind].append('')
                        if HW4GGC_WG[row[0]] in '2288X泰山' and HW4GGC_WG[row[0]]!='':
                            sheet4_lst[ind].append(HW4GGC_PQ[row[0]])
                            sheet4_lst[ind].append(HW4GGC_WG[row[0]])
                        else:
                            if HW4GGC_PQ[row[0]]=='中移3':
                                sheet4_lst[ind].append(HW4GGC_PQ[row[0]])
                                sheet4_lst[ind].append('泰山')
                            else:
                                sheet4_lst[ind].append(HW4GGC_PQ[row[0]])
                                sheet4_lst[ind].append('2288X')
                        ind += 1
                except:
                    pass
            else:
                if HW4GGC_index[str(row[1]) + '_' + str(row[2])]!=2:
                    try:
                        if HW4GGC_PQ[row[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                            try:
                                COUNT[2][HW4GGC_PQ[row[0]]] += 1
                            except:
                                COUNT[2][HW4GGC_PQ[row[0]]] = 1
                            sheet4_lst.append(list(row[:13]))
                            sheet4_lst[ind].append('')
                            if HW4GGC_WG[row[0]] in '2288X泰山':
                                sheet4_lst[ind].append(HW4GGC_PQ[row[0]])
                                sheet4_lst[ind].append(HW4GGC_WG[row[0]])
                            else:
                                if HW4GGC_PQ[row[0]] == '中移3':
                                    sheet4_lst[ind].append(HW4GGC_PQ[row[0]])
                                    sheet4_lst[ind].append('泰山')
                                else:
                                    sheet4_lst[ind].append(HW4GGC_PQ[row[0]])
                                    sheet4_lst[ind].append('2288X')
                            ind += 1
                    except:
                        pass
    print(len(sheet4_lst))
sheet5_lst=[]
def read_fourtofiveextrenal_file(file_path,file_path2):
    df = pd.read_excel(file_path)
    df2 = pd.read_excel(file_path2)
    ind=len(sheet5_lst)
    for row in df.values:
        try:
            if HW4GGC_PQ[row[2]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                try:
                    COUNT[3][HW4GGC_PQ[row[2]]] += 1
                except:
                    COUNT[3][HW4GGC_PQ[row[2]]] = 1
                sheet5_lst.append(list(row[2:14]))
                sheet5_lst[ind].append('')
                sheet5_lst[ind].append('')
                if HW4GGC_WG[row[2]] in '2288X泰山' and HW4GGC_WG[row[2]]!='':
                    sheet5_lst[ind].append(HW4GGC_PQ[row[2]])
                    sheet5_lst[ind].append(HW4GGC_WG[row[2]])
                else:
                    if HW4GGC_PQ[row[2]]=='中移3':
                        sheet5_lst[ind].append(HW4GGC_PQ[row[2]])
                        sheet5_lst[ind].append('泰山')
                    else:
                        sheet5_lst[ind].append(HW4GGC_PQ[row[2]])
                        sheet5_lst[ind].append('2288X')
                ind+=1
        except:
            pass
    print(len(sheet5_lst))
    for row in df2.values:
        if row[12]=='否':
            try:
                if HW4GGC_PQ[row[0]] in '华苏1华苏2华苏3华星1华星2欣网1欣网2欣网3欣网4中移3中移4':
                    try:
                        COUNT[3][HW4GGC_PQ[row[0]]] += 1
                    except:
                        COUNT[3][HW4GGC_PQ[row[0]]] = 1
                    sheet5_lst.append(list(row[:12]))
                    sheet5_lst[ind].append('')
                    sheet5_lst[ind].append('')
                    if HW4GGC_WG[row[0]] in '2288X泰山' and HW4GGC_WG[row[0]]!='':
                        sheet5_lst[ind].append(HW4GGC_PQ[row[0]])
                        sheet5_lst[ind].append(HW4GGC_WG[row[0]])
                    else:
                        if HW4GGC_PQ[row[0]]=='中移3':
                            sheet5_lst[ind].append(HW4GGC_PQ[row[0]])
                            sheet5_lst[ind].append('泰山')
                        else:
                            sheet5_lst[ind].append(HW4GGC_PQ[row[0]])
                            sheet5_lst[ind].append('2288X')
                    ind+=1
            except:
                pass
    print(len(sheet5_lst))
#添加数据到excel
def append_data_to_excel(file_path,sheet_name2,sheet_name3,sheet_name4,sheet_name5,data2,data3,data4,data5):
    # 加载现有的 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    # 选择要操作的工作表
    sheet2 = workbook[sheet_name2]
    sheet3 = workbook[sheet_name3]
    sheet4 = workbook[sheet_name4]
    sheet5 = workbook[sheet_name5]
    # 遍历要追加的数据列表
    for row in data2:
        # 将数据逐行添加到工作表的下一行
        sheet2.append(row)
    for row in data3:
        # 将数据逐行添加到工作表的下一行
        sheet3.append(row)
    for row in data4:
        # 将数据逐行添加到工作表的下一行
        sheet4.append(row)
    for row in data5:
        # 将数据逐行添加到工作表的下一行
        sheet5.append(row)
    # 保存更新后的 Excel 文件
    workbook.save(file_path)

def write_to_excel(file_path, sheet_name, row, column,value):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    # 获取指定工作表
    sheet = workbook[sheet_name]
    sheet.cell(row=row, column=column, value=value)
    # 在指定单元格写入数据
    #sheet.cell(row=row, column=column, value=value)
    # 保存修改后的 Excel 文件
    workbook.save(file_path)

def write_val(count:list,file_path,sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    # 获取指定工作表
    sheet = workbook['数量']
    ind=1
    for row in count:
        for li in range(11):
            cell = sheet.cell(row=li + 2, column=1)
            value = cell.value
            try:
                write_to_excel(file_path, sheet_name, li + 2, ind + 1, row[value])
            except:
                pass
        ind += 1


if __name__ == "__main__":
    day='29'
    #day=input('请输入日期')
    jieguo_url = 'G:/工作内容/25年每月例行工作/1月工作内容/冗余邻区/3月第一周/'

    fivetofivewb_path = f'{jieguo_url}结果/外部一致性/5-5外部一致性核查.xlsx'
    alx_fivetofivewb_path = f'{jieguo_url}结果/外部一致性/爱立信5-5外部一致性核查.xlsx'
    fivetofourwb_path = f'{jieguo_url}结果/外部一致性/5-4外部一致性核查.xlsx'
    alx_fivetofourwb_path = f'{jieguo_url}结果/外部一致性/爱立信5-4外部一致性核查.xlsx'
    hw_fourtofourwb_path = f'{jieguo_url}结果/外部一致性/华为4-4外部一致性核查.xlsx'
    alx_fourtofourwb_path = f'{jieguo_url}结果/外部一致性/爱立信4-4外部一致性核查.xlsx'
    fourtofivewb_path = f'{jieguo_url}结果/外部一致性/4-5外部一致性核查.xlsx'
    alx_fourtofivewb_path = f'{jieguo_url}结果/外部一致性/爱立信4-5外部一致性核查.xlsx'
    HW5GGC_path = f'{jieguo_url}解析文件/苏州华为5G工参202505{day}.csv'
    HW4GGC_path = f'{jieguo_url}解析文件/苏州华为4G工参202505{day}.csv'
    update_excel_path = f'{jieguo_url}外部不一致核查05{day}.xlsx'

    print('正在核查，请稍等......')

    read_HW5GGC_file(HW5GGC_path)
    read_HW4GGC_file(HW4GGC_path)
    #5-5
    read_fivetofiveextrenal_file(fivetofivewb_path,alx_fivetofivewb_path)
    #5-4
    read_fivetofourextrenal_file(fivetofourwb_path,'hw')
    read_fivetofourextrenal_file(alx_fivetofourwb_path,'alx')
    # 4-4
    read_fourtofourextrenal_file(hw_fourtofourwb_path,'hw')
    read_fourtofourextrenal_file(alx_fourtofourwb_path,'alx')
    # 4-5
    read_fourtofiveextrenal_file(fourtofivewb_path,alx_fourtofivewb_path)

    append_data_to_excel(update_excel_path, '5-5外部不一致', '5-4外部不一致', '4-4外部不一致','4-5外部不一致', sheet2_lst, sheet3_lst, sheet4_lst,sheet5_lst)
    write_val(COUNT,update_excel_path,'数量')