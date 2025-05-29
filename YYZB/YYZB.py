import openpyxl
import os
import zipfile
import csv
from datetime import datetime

ind = 1
YMD='20250'

def list_files_recursively(path):
    """递归列出路径下的所有文件"""
    file_list = []
    for root, dirs, files in os.walk(path):
        for file in files:
            file_path = os.path.join(root, file)
            file_list.append(file_path)
    return file_list

def append_to_excel(file_path, row_data,index, start_col=2):

    if '分厂商统计 -3天' in file_path:
        # 加载工作簿
        wb = openpyxl.load_workbook(file_path)
        ws6 = wb['定制EPSFB_5G-厂商']
        ws4 = wb['定制EPSFB_4G-厂商']
        ws2 = wb['定制VoNR-厂商']
        ws15=wb['VoLTE-厂商']

        if index == 1:
            # 在最后一行之后追加
            new_row = ws2.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws2.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws2.cell(row=new_row, column=col, value=float(value))
                except:
                    ws2.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws2.cell(row=new_row, column=1).value = first_col_value
        if index == 4:
            # 在最后一行之后追加
            new_row = ws4.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws4.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws4.cell(row=new_row, column=col, value=float(value))
                except:
                    ws4.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws4.cell(row=new_row, column=1).value = first_col_value
        if index == 6:
            # 在最后一行之后追加
            new_row = ws6.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws6.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws6.cell(row=new_row, column=col, value=float(value))
                except:
                    ws6.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws6.cell(row=new_row, column=1).value = first_col_value
        if index == 15:
            # 在最后一行之后追加
            new_row = ws15.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws15.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws15.cell(row=new_row, column=col, value=float(value))
                except:
                    ws15.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws15.cell(row=new_row, column=1).value = first_col_value

    if 'SEQ平台指标多天' in file_path:
        # 加载工作簿
        wb = openpyxl.load_workbook(file_path)
        ws2 = wb['定制VONR']
        ws3=wb['定制EPSFB4G']
        ws5=wb['定制EPSFB5G']
        ws7=wb['语音接入']
        ws8=wb['视频接入']
        ws9=wb['语音保持']
        ws10=wb['视频保持']
        ws11=wb['切换']
        ws12=wb['通话质量']
        ws13=wb['注册']
        ws14=wb['用户数']

        if index == 2:
            # 在最后一行之后追加
            new_row = ws2.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws2.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws2.cell(row=new_row, column=col, value=float(value))
                except:
                    ws2.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws2.cell(row=new_row, column=1).value = first_col_value
        if index == 3:
            # 在最后一行之后追加
            new_row = ws3.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws3.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws3.cell(row=new_row, column=col, value=float(value))
                except:
                    ws3.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws3.cell(row=new_row, column=1).value = first_col_value
        if index == 5:
            # 在最后一行之后追加
            new_row = ws5.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws5.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws5.cell(row=new_row, column=col, value=float(value))
                except:
                    ws5.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws5.cell(row=new_row, column=1).value = first_col_value
        if index == 7:
            # 在最后一行之后追加
            new_row = ws7.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws7.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws7.cell(row=new_row, column=col, value=float(value))
                except:
                    ws7.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws7.cell(row=new_row, column=1).value = first_col_value
        if index == 8:
            # 在最后一行之后追加
            new_row = ws8.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws8.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws8.cell(row=new_row, column=col, value=float(value))
                except:
                    ws8.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws8.cell(row=new_row, column=1).value = first_col_value
        if index == 9:
            # 在最后一行之后追加
            new_row = ws9.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws9.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws9.cell(row=new_row, column=col, value=float(value))
                except:
                    ws9.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws9.cell(row=new_row, column=1).value = first_col_value
        if index == 10:
            # 在最后一行之后追加
            new_row = ws10.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws10.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws10.cell(row=new_row, column=col, value=float(value))
                except:
                    ws10.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws10.cell(row=new_row, column=1).value = first_col_value
        if index == 11:
            # 在最后一行之后追加
            new_row = ws11.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws11.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws11.cell(row=new_row, column=col, value=float(value))
                except:
                    ws11.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws11.cell(row=new_row, column=1).value = first_col_value
        if index == 12:
            # 在最后一行之后追加
            new_row = ws12.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws12.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws12.cell(row=new_row, column=col, value=float(value))
                except:
                    ws12.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws12.cell(row=new_row, column=1).value = first_col_value
        if index == 13:
            # 在最后一行之后追加
            new_row = ws13.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws13.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                        value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws13.cell(row=new_row, column=col, value=float(value))
                except:
                    ws13.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws13.cell(row=new_row, column=1).value = first_col_value
        if index == 14:
            # 在最后一行之后追加
            new_row = ws14.max_row + 1
            # 保留第一列的公式（如果有）
            first_col_value = ws14.cell(row=new_row, column=1).value
            # 写入数据（从start_col列开始）
            for col, value in enumerate(row_data, start=start_col):
                if YMD in value:
                    try:
                      value = datetime.strptime(value.split(' ')[0], "%Y-%m-%d").strftime('%Y/%m/%d')
                    except:
                        value = datetime.strptime(value.split(' ')[0], "%Y/%m/%d")
                try:
                    ws14.cell(row=new_row, column=col, value=float(value))
                except:
                    ws14.cell(row=new_row, column=col, value=value)

            # 恢复第一列的公式（如果有）
            if first_col_value and str(first_col_value).startswith('='):
                ws14.cell(row=new_row, column=1).value = first_col_value
    wb.save(file_path)

def unzip_file(zip_path,num, extract_to=None):
    global ind
    # 确保ZIP文件存在
    if not os.path.exists(zip_path):
        raise FileNotFoundError(f"ZIP文件不存在: {zip_path}")

    # 获取ZIP文件所在目录作为解压路径(如果不指定解压路径)
    if extract_to is None:
        extract_to = os.path.dirname(os.path.abspath(zip_path))

    # 创建解压目录(如果不存在)
    os.makedirs(extract_to, exist_ok=True)

    # 解压文件
    if 'VoNR性能分析_全量样本数据' in zip_path:
        ind+=1
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            extract_to=extract_to+'\\'+'VoNR性能分析_全量样本数据'+str(num)
            zip_ref.extractall(extract_to)
            return extract_to
    else:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_to)

    file_path=zip_path.replace('zip','csv')
    return file_path


# 使用示例
if __name__ == "__main__":
    seq_source_file='G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/每日数据/5月/09/SEQ/'
    file_lst=list_files_recursively(seq_source_file)
    VONR=[]
    EFPSFB=[]
    VONRXN=[]
    DT=''
    RTP=''
    ZC=''
    YH=''
    VOLTE_CJ=''
    for file in file_lst:
        if file.endswith('.zip'):
            if 'VOLTE小区指标' not in file:
                file_path=unzip_file(file,ind)
                if '江苏移动定制VoNR语音视频通话质量报表' in file_path:
                    VONR.append(file_path)
                if 'VoNR性能分析_全量样本数据' in file_path:
                    VONRXN.append(file_path)
                if '江苏移动定制EPSFB' in file_path:
                    EFPSFB.append(file_path)
                if 'DT' in file_path:
                    DT=file_path
                if 'RTP' in file_path:
                    RTP=file_path
                if 'ZC' in file_path:
                    ZC=file_path
                    with open(ZC, 'r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            if '2025' in row[0]:
                                append_to_excel(
                                                    "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx",
                                                    row, 13)
                if 'YH' in file_path:
                    YH=file_path
                    with open(YH, 'r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            if '2025' in row[0]:
                                append_to_excel(
                                                    "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx",
                                                    row, 14)
        elif 'VOLTE厂家指标' in file:
            VOLTE_CJ=file
            with open(VOLTE_CJ, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    if '2025' in row[1]:
                        append_to_excel(
                            "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/分厂商统计 -3天 - 副本.xlsx",
                            row, 15)

    for li in VONR:
        with open(li, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                if row[2]=='区县':
                    break
                elif row[2]=='5G IMS初始注册成功率':
                    with open(li, 'r', encoding='utf-8') as f1:
                        r1 = csv.reader(f1)
                        for row in r1:
                            if row[2] != '5G IMS初始注册成功率':
                                append_to_excel("G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx", row,2)
                    break
                elif row[2]=='厂商':
                    with open(li, 'r', encoding='utf-8') as f1:
                        r1 = csv.reader(f1)
                        for row in r1:
                            if row[2] !='厂商':
                                append_to_excel("G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/分厂商统计 -3天 - 副本.xlsx", row,1)
                    break

    for li in EFPSFB:
        if '江苏移动定制EPSFB_4G' in li:
            with open(li, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    if row[2] == '厂商':
                        with open(li, 'r', encoding='utf-8') as f:
                            reader = csv.reader(f)
                            for row in reader:
                                if row[2] != '厂商':
                                    append_to_excel(
                                        "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/分厂商统计 -3天 - 副本.xlsx",
                                        row, 4)
                    if row[2] == 'EPS FB上行RTP丢包率':
                        with open(li, 'r', encoding='utf-8') as f:
                            reader = csv.reader(f)
                            for row in reader:
                                if row[2] != 'EPS FB上行RTP丢包率':
                                    append_to_excel(
                                        "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx",
                                        row, 3)
        if '江苏移动定制EPSFB_5G' in li:
            with open(li, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    if row[2] == '厂商':
                        with open(li, 'r', encoding='utf-8') as f:
                            reader = csv.reader(f)
                            for row in reader:
                                if row[2] != '厂商':
                                    append_to_excel(
                                        "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/分厂商统计 -3天 - 副本.xlsx",
                                        row, 6)
                    if row[2] == 'EPSFB回落请求次数':
                        with open(li, 'r', encoding='utf-8') as f:
                            reader = csv.reader(f)
                            for row in reader:
                                if row[2] != 'EPSFB回落请求次数':
                                    append_to_excel(
                                        "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx",
                                        row, 5)

    for li in VONRXN:
        file_path=li+'\\报表详情.csv'
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                if '2025' in row[0] and len(row[0])<11:
                    if len(row)==39:
                        if float(row[2])<96:
                            append_to_excel(
                                "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx",
                                row, 7)
                        else:
                            append_to_excel(
                                "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx",
                                row, 8)
                    if len(row)==24:
                        if float(row[3])>1500000:
                            append_to_excel(
                                "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx",
                                row, 9)
                        else:
                            append_to_excel(
                                "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx",
                                row, 10)
                    if len(row)==20:
                        append_to_excel(
                            "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx",
                            row, 11)
                    if len(row)==125:
                        append_to_excel(
                            "G:/工作内容/25年每月例行工作/1月工作内容/语音指标统计/模版/模板3.0/SEQ平台指标多天 - 副本 - 副本.xlsx",
                            row, 12)